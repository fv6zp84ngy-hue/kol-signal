from __future__ import annotations

import tempfile
import unittest
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

from kol_signal.core import (
    aggregate_group,
    identity_resolution,
    load_records,
    normalize_handle,
    normalize_url,
    parse_campaign,
    parse_datetime,
    parse_number,
    run_pipeline,
    score_creator,
)


ROOT = Path(__file__).resolve().parents[1]
FIXTURES = ROOT / "fixtures"
REFERENCE_TIME = datetime(2026, 7, 29, tzinfo=timezone.utc)


class NormalizationTests(unittest.TestCase):
    def test_compact_numbers_percentages_and_invalid_values(self) -> None:
        self.assertEqual(parse_number("128K"), (128000, None))
        self.assertEqual(parse_number("1.2M"), (1200000, None))
        self.assertEqual(parse_number("3.5%", percentage=True), (0.035, None))
        self.assertEqual(parse_number("many"), (None, "many"))
        self.assertEqual(parse_datetime("2026-13-40"), (None, "2026-13-40"))

    def test_handle_and_url_are_normalized(self) -> None:
        self.assertEqual(normalize_handle(" @PetOrbitDaily "), "petorbitdaily")
        self.assertEqual(
            normalize_url("https://Example.com/petorbitdaily/?utm_source=wave&ref=list"),
            "https://example.com/petorbitdaily",
        )


class IdentityAndQualityTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.all_records, cls.mapping = load_records(
            [
                FIXTURES / "waveinflu_creators.xlsx",
                FIXTURES / "nox_creators.csv",
                FIXTURES / "manual_creators.xlsx",
            ]
        )
        cls.groups, cls.review_items = identity_resolution(cls.all_records)
        cls.creators = [aggregate_group(group, REFERENCE_TIME) for group in cls.groups]

    def creator_for_record(self, record_id: str):
        return next(
            creator for creator in self.creators if record_id in creator.source_record_ids
        )

    def test_three_source_fixture_matches_expected_merge_counts(self) -> None:
        self.assertEqual(len(self.all_records), 120)
        self.assertEqual(len(self.groups), 85)
        merged_groups = [group for group in self.groups if len(group) > 1]
        self.assertEqual(len(merged_groups), 25)
        self.assertEqual(sum(len(group) for group in merged_groups), 60)

    def test_three_source_duplicate_is_one_creator(self) -> None:
        creator = self.creator_for_record("WAVE-001")
        self.assertEqual(
            creator.source_record_ids,
            ["MAN-001", "NOX-001", "WAVE-001"],
        )

    def test_same_handle_with_different_platform_id_is_not_merged(self) -> None:
        wave_creator = self.creator_for_record("WAVE-021")
        nox_creator = self.creator_for_record("NOX-021")
        self.assertNotEqual(wave_creator.creator_id, nox_creator.creator_id)
        review_text = "\n".join(str(item) for item in self.review_items)
        self.assertIn("WAVE-021", review_text)
        self.assertIn("NOX-021", review_text)

    def test_shared_agency_email_is_not_an_identity_key(self) -> None:
        wave_creator = self.creator_for_record("WAVE-022")
        nox_creator = self.creator_for_record("NOX-022")
        self.assertNotEqual(wave_creator.creator_id, nox_creator.creator_id)
        review_text = "\n".join(str(item) for item in self.review_items)
        self.assertIn("shared agency email", review_text)

    def test_high_follower_conflict_and_unknown_observation_time(self) -> None:
        creator = self.creator_for_record("WAVE-002")
        self.assertIn(
            ("followers", "High"),
            [(item.field, item.level) for item in creator.conflicts],
        )
        nox_010 = next(record for record in self.all_records if record.source_record_id == "NOX-010")
        single = aggregate_group([nox_010], REFERENCE_TIME)
        self.assertIn(
            ("followers", "Observed Time Unknown"),
            [(item.field, item.status) for item in single.stale_items],
        )

    def test_invalid_value_is_preserved_and_not_zero(self) -> None:
        wave_030 = next(record for record in self.all_records if record.source_record_id == "WAVE-030")
        self.assertIsNone(wave_030.followers)
        self.assertEqual(wave_030.invalid_fields["followers"], "many")

    def test_hard_filters_and_four_dimension_scores(self) -> None:
        campaign = parse_campaign(FIXTURES / "campaign.txt")
        for creator in self.creators:
            score_creator(creator, campaign, REFERENCE_TIME)
            self.assertEqual(
                set(creator.dimensions),
                {"brand_fit", "commercial_readiness", "contactability", "data_confidence"},
            )
            if creator.action_level == "Priority":
                self.assertGreaterEqual(len(creator.why_contact), 2)
        all_exclusions = {
            reason for creator in self.creators for reason in creator.exclusion_reasons
        }
        self.assertTrue(
            {
                "EXCLUDED_MARKET",
                "EXCLUDED_LANGUAGE",
                "EXCLUDED_FOLLOWER_RANGE",
                "EXCLUDED_INACTIVE",
                "EXCLUDED_NO_CONTACT",
            }.issubset(all_exclusions)
        )


class EndToEndTests(unittest.TestCase):
    def test_target_command_generates_required_run_artifacts(self) -> None:
        with tempfile.TemporaryDirectory() as temporary_directory:
            output_base = Path(temporary_directory) / "runs"
            result = run_pipeline(
                input_paths=[
                    FIXTURES / "waveinflu_creators.xlsx",
                    FIXTURES / "nox_creators.csv",
                ],
                brief_path=FIXTURES / "campaign.txt",
                output_base=output_base,
            )
            self.assertEqual(result.raw_record_count, 80)
            self.assertEqual(result.creator_count, 65)
            self.assertEqual(result.review_count, 2)
            self.assertEqual(
                {path.name for path in result.run_dir.iterdir()},
                {
                    "data_audit.xlsx",
                    "creator_shortlist.xlsx",
                    "merge_review.xlsx",
                    "feedback_template.xlsx",
                    "report.md",
                    "campaign.json",
                    "manifest.json",
                },
            )

            audit = load_workbook(result.run_dir / "data_audit.xlsx", read_only=True)
            self.assertEqual(
                audit.sheetnames,
                [
                    "Summary",
                    "Source Comparison",
                    "Field Coverage",
                    "Conflicts",
                    "Stale Data",
                    "Duplicate Candidates",
                ],
            )
            audit.close()

            shortlist = load_workbook(result.run_dir / "creator_shortlist.xlsx", read_only=True)
            self.assertEqual(shortlist["Shortlist"].max_row, 66)
            shortlist.close()

            review = load_workbook(result.run_dir / "merge_review.xlsx", read_only=True)
            self.assertEqual(review["Review Queue"].max_row, 3)
            review.close()

            feedback = load_workbook(
                result.run_dir / "feedback_template.xlsx",
                read_only=True,
            )
            self.assertEqual(feedback["Feedback"].max_row, 66)
            feedback.close()

            report = (result.run_dir / "report.md").read_text(encoding="utf-8")
            self.assertIn("原始记录：80", report)
            self.assertIn("去重后达人：65", report)


if __name__ == "__main__":
    unittest.main()
