from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from kol_signal.core import prepare_input_mappings, review_run, run_pipeline
from kol_signal.feedback import calculate_metrics, import_feedback


ROOT = Path(__file__).resolve().parents[1]
FIXTURES = ROOT / "fixtures"


class ReviewAndFeedbackTests(unittest.TestCase):
    def create_run(self, output_base: Path):
        inputs = [
            FIXTURES / "waveinflu_creators.xlsx",
            FIXTURES / "nox_creators.csv",
        ]
        plans = prepare_input_mappings(inputs)
        return run_pipeline(
            input_paths=inputs,
            brief_path=FIXTURES / "campaign.txt",
            output_base=output_base,
            mapping_plans=plans,
        )

    def test_review_recomputes_run_and_feedback_metrics(self) -> None:
        with tempfile.TemporaryDirectory() as temporary_directory:
            runs_dir = Path(temporary_directory) / "runs"
            initial = self.create_run(runs_dir)
            review_path = initial.run_dir / "merge_review.xlsx"
            review_workbook = load_workbook(review_path)
            review_sheet = review_workbook["Review Queue"]
            self.assertEqual(review_sheet.max_row, 3)
            review_sheet["F2"] = "merge"
            review_sheet["G2"] = "Confirmed as one creator."
            review_sheet["F3"] = "keep_separate"
            review_sheet["G3"] = "Shared agency mailbox only."
            review_workbook.save(review_path)
            review_workbook.close()

            reviewed = review_run(
                run_id=initial.run_id,
                review_input=review_path,
                runs_dir=runs_dir,
            )
            self.assertEqual(reviewed.creator_count, 64)
            self.assertEqual(reviewed.review_count, 0)

            manifest = json.loads(
                (reviewed.run_dir / "manifest.json").read_text(encoding="utf-8")
            )
            self.assertEqual(manifest["review_revision"], 1)
            self.assertEqual(
                [item["user_decision"] for item in manifest["review_decisions"]],
                ["merge", "keep_separate"],
            )

            feedback_path = reviewed.run_dir / "feedback_template.xlsx"
            feedback_workbook = load_workbook(feedback_path)
            feedback_sheet = feedback_workbook["Feedback"]
            values = [
                # merge, shortlist, contact, contacted, delivered, bounced, replied, positive
                (True, True, True, True, True, False, True, True, "推荐合理"),
                (True, False, False, True, False, True, False, False, "邮箱退信"),
                (False, True, None, False, False, False, False, False, "需要补查"),
            ]
            for row_number, row_values in enumerate(values, start=2):
                for column, value in enumerate(row_values[:8], start=7):
                    feedback_sheet.cell(row_number, column).value = value
                feedback_sheet.cell(row_number, 15).value = row_values[8]
            feedback_workbook.save(feedback_path)
            feedback_workbook.close()

            result = import_feedback(
                run_id=initial.run_id,
                feedback_input=feedback_path,
                runs_dir=runs_dir,
            )
            metrics = {item["key"]: item["value"] for item in result.metrics}
            self.assertAlmostEqual(metrics["merge_accuracy"], 2 / 3)
            self.assertAlmostEqual(metrics["shortlist_acceptance_rate"], 2 / 3)
            self.assertAlmostEqual(metrics["contact_accuracy"], 1 / 2)
            self.assertEqual(metrics["actually_contacted"], 2)
            self.assertAlmostEqual(metrics["delivery_rate"], 1 / 2)
            self.assertEqual(metrics["reply_rate"], 1.0)
            self.assertEqual(metrics["positive_reply_rate"], 1.0)
            self.assertEqual(result.note_count, 3)
            self.assertEqual(result.warning_count, 0)

            report = load_workbook(result.run_dir / "feedback_report.xlsx", read_only=True)
            self.assertEqual(
                report.sheetnames,
                [
                    "Summary",
                    "Imported Feedback",
                    "Text Feedback",
                    "Validation Warnings",
                ],
            )
            self.assertAlmostEqual(report["Summary"]["B4"].value, 2 / 3)
            report.close()
            markdown = (result.run_dir / "feedback_report.md").read_text(
                encoding="utf-8"
            )
            self.assertIn("合并准确率：66.7%（2/3）", markdown)
            self.assertIn("推荐合理", markdown)

    def test_zero_denominators_are_null_not_zero_percent(self) -> None:
        metrics = {
            item["key"]: item["value"]
            for item in calculate_metrics([])
        }
        self.assertIsNone(metrics["merge_accuracy"])
        self.assertIsNone(metrics["shortlist_acceptance_rate"])
        self.assertIsNone(metrics["contact_accuracy"])
        self.assertEqual(metrics["actually_contacted"], 0)
        self.assertIsNone(metrics["delivery_rate"])
        self.assertIsNone(metrics["reply_rate"])
        self.assertIsNone(metrics["positive_reply_rate"])


if __name__ == "__main__":
    unittest.main()
