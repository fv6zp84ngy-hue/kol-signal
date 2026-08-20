from __future__ import annotations

import base64
import io
import json
import tempfile
import unittest
from contextlib import redirect_stderr, redirect_stdout
from pathlib import Path
from unittest import mock

from openpyxl import Workbook, load_workbook

from kol_signal.cli import main
from kol_signal.core import (
    PipelineError,
    load_records,
    prepare_input_mappings,
    read_tabular,
    run_pipeline,
)
from kol_signal.reporting import safe_excel_value


ROOT = Path(__file__).resolve().parents[1]
FIXTURES = ROOT / "fixtures"
SECURITY_FIXTURES = FIXTURES / "security"


class ExcelFormulaSafetyTests(unittest.TestCase):
    def test_dangerous_text_is_always_plain_excel_text(self) -> None:
        cases = json.loads(
            (SECURITY_FIXTURES / "golden_cases.json").read_text(encoding="utf-8")
        )
        for value in cases["dangerous_excel_prefixes"]:
            with self.subTest(value=repr(value)):
                protected = safe_excel_value(value)
                self.assertIsInstance(protected, str)
                self.assertTrue(protected.startswith("'"))

        self.assertEqual(safe_excel_value(10), 10)
        self.assertEqual(safe_excel_value(-1), -1)

    def test_all_four_output_workbooks_contain_no_formula_cells(self) -> None:
        with tempfile.TemporaryDirectory() as temporary_directory:
            result = run_pipeline(
                input_paths=[SECURITY_FIXTURES / "dangerous_values.csv"],
                brief_path=FIXTURES / "campaign.txt",
                output_base=Path(temporary_directory) / "runs",
            )
            workbook_names = {
                "data_audit.xlsx",
                "creator_shortlist.xlsx",
                "merge_review.xlsx",
                "feedback_template.xlsx",
            }
            for name in workbook_names:
                with self.subTest(workbook=name):
                    workbook = load_workbook(result.run_dir / name, data_only=False)
                    cells = [
                        cell
                        for worksheet in workbook.worksheets
                        for row in worksheet.iter_rows()
                        for cell in row
                        if cell.value is not None
                    ]
                    self.assertFalse(
                        [cell.coordinate for cell in cells if cell.data_type == "f"]
                    )
                    self.assertTrue(
                        any(
                            isinstance(cell.value, str) and cell.value.startswith("'")
                            for cell in cells
                        )
                    )
                    workbook.close()


class InputSafetyTests(unittest.TestCase):
    def test_corrupt_xlsx_has_stable_cli_error_without_traceback(self) -> None:
        stdout = io.StringIO()
        stderr = io.StringIO()
        with redirect_stdout(stdout), redirect_stderr(stderr):
            exit_code = main(
                [
                    "mapping-preview",
                    "--input",
                    str(SECURITY_FIXTURES / "corrupt.xlsx"),
                ]
            )
        self.assertEqual(exit_code, 2)
        self.assertIn("Could not read XLSX", stderr.getvalue())
        self.assertNotIn("Traceback", stderr.getvalue())
        self.assertEqual(stdout.getvalue(), "")

    def test_non_utf8_csv_has_clear_encoding_error_without_traceback(self) -> None:
        raw = base64.b64decode(
            (SECURITY_FIXTURES / "invalid_utf8.csv.base64").read_text(
                encoding="ascii"
            )
        )
        with tempfile.TemporaryDirectory() as temporary_directory:
            path = Path(temporary_directory) / "invalid.csv"
            path.write_bytes(raw)
            stderr = io.StringIO()
            with redirect_stderr(stderr):
                exit_code = main(["mapping-preview", "--input", str(path)])
        self.assertEqual(exit_code, 2)
        self.assertIn("Only UTF-8 or UTF-8-SIG CSV is supported", stderr.getvalue())
        self.assertNotIn("Traceback", stderr.getvalue())

    def test_empty_and_duplicate_headers_fail_explicitly(self) -> None:
        with self.assertRaisesRegex(PipelineError, "empty header"):
            read_tabular(SECURITY_FIXTURES / "empty_header.csv")
        with self.assertRaisesRegex(PipelineError, "duplicate header"):
            read_tabular(SECURITY_FIXTURES / "duplicate_header.csv")

    def test_multiple_non_empty_worksheets_require_explicit_selection(self) -> None:
        with tempfile.TemporaryDirectory() as temporary_directory:
            path = Path(temporary_directory) / "multiple.xlsx"
            workbook = Workbook()
            first = workbook.active
            first.title = "Creators A"
            first.append(["Platform", "Handle"])
            first.append(["TikTok", "@fixture_a"])
            second = workbook.create_sheet("Creators B")
            second.append(["Platform", "Handle"])
            second.append(["TikTok", "@fixture_b"])
            workbook.save(path)
            workbook.close()

            with self.assertRaisesRegex(PipelineError, "multiple non-empty worksheets"):
                read_tabular(path)
            headers, rows = read_tabular(path, sheet_name="Creators B")
            self.assertEqual(headers, ["Platform", "Handle"])
            self.assertEqual(rows[0]["Handle"], "@fixture_b")

    def test_output_permission_error_does_not_overwrite_existing_run(self) -> None:
        input_path = FIXTURES / "nox_creators.csv"
        plans = prepare_input_mappings([input_path])
        with tempfile.TemporaryDirectory() as temporary_directory:
            output_base = Path(temporary_directory) / "runs"
            existing = output_base / "run_fixed"
            existing.mkdir(parents=True)
            sentinel = existing / "keep.txt"
            sentinel.write_text("original", encoding="utf-8")

            with (
                mock.patch("kol_signal.core.make_run_id", return_value="run_fixed"),
                mock.patch.object(
                    Path,
                    "mkdir",
                    side_effect=PermissionError("permission denied"),
                ),
                self.assertRaisesRegex(PipelineError, "Cannot create output directory"),
            ):
                run_pipeline(
                    input_paths=[input_path],
                    brief_path=FIXTURES / "campaign.txt",
                    output_base=output_base,
                    mapping_plans=plans,
                )

            self.assertEqual(sentinel.read_text(encoding="utf-8"), "original")
            self.assertEqual(
                sorted(path.name for path in output_base.iterdir()),
                ["run_fixed"],
            )

    def test_invalid_row_does_not_block_other_valid_rows(self) -> None:
        with tempfile.TemporaryDirectory() as temporary_directory:
            path = Path(temporary_directory) / "rows.csv"
            path.write_text(
                "Nox ID,Channel,Channel ID,Handle,URL,Total Followers\n"
                "GOOD-1,Tik Tok,id-good,@good,,128K\n"
                "BAD-2,Tik Tok,id-bad,@bad,,not-a-number\n",
                encoding="utf-8",
            )
            records, _ = load_records([path])
        self.assertEqual(len(records), 2)
        invalid = next(record for record in records if record.source_record_id == "BAD-2")
        self.assertIsNone(invalid.followers)
        self.assertEqual(invalid.invalid_fields["followers"], "not-a-number")


if __name__ == "__main__":
    unittest.main()
