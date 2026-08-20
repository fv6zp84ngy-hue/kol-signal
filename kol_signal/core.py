from __future__ import annotations

import csv
import hashlib
import json
import math
import re
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime, time, timezone
from pathlib import Path
from typing import Any, Iterable
from urllib.parse import parse_qsl, urlencode, urlsplit, urlunsplit
from zipfile import BadZipFile
from xml.etree.ElementTree import ParseError

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from .mapping import (
    CANONICAL_FIELDS,
    InputMappingPlan,
    confirm_ambiguous_columns,
    create_mapping_plan,
    save_mapping_config,
    validate_mapping,
)
from .reporting import (
    write_creator_shortlist,
    write_data_audit,
    write_feedback_template,
    write_merge_review,
    write_report,
)


class PipelineError(RuntimeError):
    """A user-actionable pipeline error."""

    default_code = "KS_PIPELINE_001"

    def __init__(self, message: str, *, code: str | None = None) -> None:
        super().__init__(message)
        self.code = code or self.default_code


class OutputError(PipelineError):
    """A user-actionable output failure that maps to CLI exit code 5."""

    default_code = "KS_OUTPUT_001"


CRITICAL_FIELDS = (
    "platform",
    "handle",
    "profile_url",
    "followers",
    "average_views",
    "engagement_rate",
    "country",
    "language",
    "email",
    "latest_post_at",
    "latest_sponsored_post_at",
)

TTL_DAYS = {
    "latest_post_at": 7,
    "followers": 30,
    "average_views": 30,
    "email": 90,
    "latest_sponsored_post_at": 45,
}

SOURCE_PRIORITY = {"manual": 3, "waveinflu": 2, "nox": 1, "generic": 0}
CAMPAIGN_PARSER_VERSION = "1"

PLATFORM_ALIASES = {
    "tiktok": "tiktok",
    "tik tok": "tiktok",
    "instagram": "instagram",
    "ig": "instagram",
    "youtube": "youtube",
    "yt": "youtube",
}

COUNTRY_ALIASES = {
    "united states": "US",
    "united states of america": "US",
    "usa": "US",
    "us": "US",
    "canada": "CA",
    "ca": "CA",
    "brazil": "BR",
    "br": "BR",
    "japan": "JP",
    "jp": "JP",
    "germany": "DE",
    "de": "DE",
}

LANGUAGE_ALIASES = {
    "english": "en",
    "en": "en",
    "spanish": "es",
    "es": "es",
    "portuguese": "pt",
    "pt": "pt",
    "japanese": "ja",
    "ja": "ja",
    "german": "de",
    "de": "de",
}

TRACKING_QUERY_KEYS = {"ref", "fbclid", "gclid", "source"}


@dataclass(slots=True)
class Campaign:
    brand: str
    markets: tuple[str, ...]
    languages: tuple[str, ...]
    platforms: tuple[str, ...]
    follower_min: int | None
    follower_max: int | None
    latest_post_max_age_days: int | None
    require_contact_path: bool | None
    raw_text: str
    blocklist: tuple[str, ...] = ()
    topics: tuple[str, ...] = ()


@dataclass(slots=True)
class CampaignParseResult:
    parsed_campaign: Campaign
    recognized_conditions: tuple[dict[str, Any], ...]
    unrecognized_conditions: tuple[dict[str, str], ...]
    conflicting_conditions: tuple[dict[str, Any], ...]
    warnings: tuple[str, ...]
    parser_version: str = CAMPAIGN_PARSER_VERSION

    @property
    def blocking_unrecognized_conditions(self) -> tuple[dict[str, str], ...]:
        return tuple(
            item
            for item in self.unrecognized_conditions
            if item.get("severity") == "blocking"
        )

    @property
    def requires_confirmation(self) -> bool:
        return bool(self.unrecognized_conditions)

    @property
    def requires_resolution(self) -> bool:
        return bool(self.conflicting_conditions)

    def to_dict(self) -> dict[str, Any]:
        campaign = self.parsed_campaign
        return {
            "parser_version": self.parser_version,
            "parsed_campaign": {
                "brand": campaign.brand,
                "markets": list(campaign.markets),
                "languages": list(campaign.languages),
                "platforms": list(campaign.platforms),
                "follower_min": campaign.follower_min,
                "follower_max": campaign.follower_max,
                "latest_post_max_age_days": campaign.latest_post_max_age_days,
                "require_contact_path": campaign.require_contact_path,
                "blocklist": list(campaign.blocklist),
                "topics": list(campaign.topics),
            },
            "recognized_conditions": [
                dict(item) for item in self.recognized_conditions
            ],
            "unrecognized_conditions": [
                dict(item) for item in self.unrecognized_conditions
            ],
            "conflicting_conditions": [
                dict(item) for item in self.conflicting_conditions
            ],
            "warnings": list(self.warnings),
        }

    @classmethod
    def from_dict(
        cls,
        payload: dict[str, Any],
        *,
        raw_text: str,
    ) -> CampaignParseResult:
        if payload.get("parser_version") != CAMPAIGN_PARSER_VERSION:
            raise PipelineError(
                "Campaign config parser_version does not match this application."
            )
        values = payload.get("parsed_campaign")
        if not isinstance(values, dict):
            raise PipelineError("Campaign config has no parsed_campaign object.")
        campaign = Campaign(
            brand=str(values.get("brand") or "Unspecified Brand"),
            markets=tuple(str(item) for item in values.get("markets", [])),
            languages=tuple(str(item) for item in values.get("languages", [])),
            platforms=tuple(str(item) for item in values.get("platforms", [])),
            follower_min=values.get("follower_min"),
            follower_max=values.get("follower_max"),
            latest_post_max_age_days=values.get("latest_post_max_age_days"),
            require_contact_path=values.get("require_contact_path"),
            raw_text=raw_text,
            blocklist=tuple(str(item) for item in values.get("blocklist", [])),
            topics=tuple(str(item) for item in values.get("topics", [])),
        )
        return cls(
            parsed_campaign=campaign,
            recognized_conditions=tuple(
                dict(item) for item in payload.get("recognized_conditions", [])
            ),
            unrecognized_conditions=tuple(
                dict(item) for item in payload.get("unrecognized_conditions", [])
            ),
            conflicting_conditions=tuple(
                dict(item) for item in payload.get("conflicting_conditions", [])
            ),
            warnings=tuple(str(item) for item in payload.get("warnings", [])),
            parser_version=str(payload["parser_version"]),
        )


@dataclass(slots=True)
class NormalizedRecord:
    source: str
    source_path: str
    row_number: int
    source_record_id: str
    platform: str | None
    platform_creator_id: str | None
    handle: str | None
    profile_url: str | None
    display_name: str | None
    followers: int | None
    average_views: int | None
    engagement_rate: float | None
    country: str | None
    language: str | None
    email: str | None
    email_role: str | None
    latest_post_at: datetime | None
    latest_sponsored_post_at: datetime | None
    observed_at: datetime | None
    email_observed_at: datetime | None
    is_estimated: bool
    raw_payload: dict[str, Any]
    invalid_fields: dict[str, str] = field(default_factory=dict)


@dataclass(slots=True)
class Conflict:
    creator_id: str
    field: str
    level: str
    values: str
    sources: str
    reason: str


@dataclass(slots=True)
class StaleItem:
    creator_id: str
    field: str
    selected_value: Any
    source: str
    observed_at: datetime | None
    ttl_days: int
    status: str


@dataclass(slots=True)
class CreatorAggregate:
    creator_id: str
    records: list[NormalizedRecord]
    selected: dict[str, Any]
    selected_meta: dict[str, dict[str, Any]]
    conflicts: list[Conflict]
    stale_items: list[StaleItem]
    missing_fields: list[str]
    invalid_fields: list[str]
    merge_basis: str
    review_required: bool = False
    action_level: str = "Hold"
    reference_score: float = 0.0
    dimensions: dict[str, float] = field(default_factory=dict)
    why_contact: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    refresh_recommendation: list[str] = field(default_factory=list)
    exclusion_reasons: list[str] = field(default_factory=list)

    @property
    def source_names(self) -> list[str]:
        return sorted({record.source for record in self.records})

    @property
    def source_record_ids(self) -> list[str]:
        return sorted(record.source_record_id for record in self.records)


@dataclass(slots=True)
class PipelineResult:
    run_id: str
    run_dir: Path
    raw_record_count: int
    creator_count: int
    review_count: int
    output_files: tuple[Path, ...]


class DisjointSet:
    def __init__(self, size: int) -> None:
        self.parent = list(range(size))
        self.rank = [0] * size

    def find(self, value: int) -> int:
        while self.parent[value] != value:
            self.parent[value] = self.parent[self.parent[value]]
            value = self.parent[value]
        return value

    def union(self, left: int, right: int) -> None:
        left_root = self.find(left)
        right_root = self.find(right)
        if left_root == right_root:
            return
        if self.rank[left_root] < self.rank[right_root]:
            left_root, right_root = right_root, left_root
        self.parent[right_root] = left_root
        if self.rank[left_root] == self.rank[right_root]:
            self.rank[left_root] += 1


def clean_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def normalize_platform(value: Any) -> str | None:
    text = clean_text(value)
    if not text:
        return None
    return PLATFORM_ALIASES.get(text.lower(), text.lower())


def normalize_handle(value: Any) -> str | None:
    text = clean_text(value)
    if not text:
        return None
    return text.lstrip("@").strip().lower() or None


def normalize_url(value: Any) -> str | None:
    text = clean_text(value)
    if not text:
        return None
    try:
        parsed = urlsplit(text)
    except ValueError:
        return None
    if not parsed.netloc:
        return text.rstrip("/")
    filtered_query = []
    for key, query_value in parse_qsl(parsed.query, keep_blank_values=True):
        key_lower = key.lower()
        if key_lower.startswith("utm_") or key_lower in TRACKING_QUERY_KEYS:
            continue
        filtered_query.append((key, query_value))
    return urlunsplit(
        (
            parsed.scheme.lower() or "https",
            parsed.netloc.lower(),
            parsed.path.rstrip("/"),
            urlencode(filtered_query),
            "",
        )
    )


def parse_number(value: Any, *, percentage: bool = False) -> tuple[float | int | None, str | None]:
    if value is None or value == "":
        return None, None
    if isinstance(value, bool):
        return None, str(value)
    if isinstance(value, (int, float)):
        numeric = float(value)
        if not math.isfinite(numeric) or numeric < 0:
            return None, str(value)
        if percentage:
            if numeric > 1:
                numeric /= 100
            return numeric, None
        return int(round(numeric)), None

    text = str(value).strip()
    if not text:
        return None, None
    percent_mark = text.endswith("%")
    multiplier = 1.0
    suffix = text[-1:].lower()
    if suffix == "k":
        multiplier = 1_000
        text = text[:-1]
    elif suffix == "m":
        multiplier = 1_000_000
        text = text[:-1]
    text = text.rstrip("%").replace(",", "").strip()
    try:
        numeric = float(text) * multiplier
    except ValueError:
        return None, str(value)
    if not math.isfinite(numeric) or numeric < 0:
        return None, str(value)
    if percentage:
        if percent_mark or numeric > 1:
            numeric /= 100
        if numeric > 1:
            return None, str(value)
        return numeric, None
    return int(round(numeric)), None


def parse_datetime(value: Any) -> tuple[datetime | None, str | None]:
    if value is None or value == "":
        return None, None
    if isinstance(value, datetime):
        parsed = value
    elif isinstance(value, date):
        parsed = datetime.combine(value, time.min)
    else:
        text = str(value).strip()
        if not text:
            return None, None
        normalized = text.replace("Z", "+00:00")
        try:
            parsed = datetime.fromisoformat(normalized)
        except ValueError:
            try:
                parsed = datetime.strptime(text, "%Y/%m/%d")
            except ValueError:
                return None, str(value)
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=timezone.utc)
    return parsed.astimezone(timezone.utc), None


def parse_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    text = (clean_text(value) or "").lower()
    return text in {"1", "true", "yes", "y", "estimated"}


def normalize_country(value: Any) -> str | None:
    text = clean_text(value)
    if not text:
        return None
    return COUNTRY_ALIASES.get(text.lower(), text.upper())


def normalize_language(value: Any) -> str | None:
    text = clean_text(value)
    if not text:
        return None
    return LANGUAGE_ALIASES.get(text.lower(), text.lower())


def normalize_email(value: Any) -> tuple[str | None, str | None]:
    text = clean_text(value)
    if not text:
        return None, None
    normalized = text.lower()
    if not re.fullmatch(r"[^\s@]+@[^\s@]+\.[^\s@]+", normalized):
        return None, text
    return normalized, None


def validate_headers(raw_headers: Iterable[Any], path: Path) -> list[str]:
    values = list(raw_headers)
    if not values:
        raise PipelineError(f"Input has no header row: {path}")
    headers = [str(value).strip() if value is not None else "" for value in values]
    empty_columns = [str(index) for index, header in enumerate(headers, start=1) if not header]
    if empty_columns:
        raise PipelineError(
            f"Input has an empty header at column(s) {', '.join(empty_columns)}: {path}"
        )
    counts = Counter(header.casefold() for header in headers)
    duplicates = sorted(
        {header for header in headers if counts[header.casefold()] > 1},
        key=str.casefold,
    )
    if duplicates:
        raise PipelineError(
            f"Input has duplicate header(s) {', '.join(duplicates)}: {path}"
        )
    return headers


def _worksheet_has_content(worksheet) -> bool:
    return any(
        any(value not in (None, "") for value in row)
        for row in worksheet.iter_rows(values_only=True)
    )


def read_tabular(
    path: Path,
    *,
    sheet_name: str | None = None,
) -> tuple[list[str], list[dict[str, Any]]]:
    if not path.exists():
        raise PipelineError(f"Input file does not exist: {path}")
    if not path.is_file():
        raise PipelineError(f"Input path is not a file: {path}")
    suffix = path.suffix.lower()
    if suffix == ".csv":
        if sheet_name:
            raise PipelineError(f"--sheet can only be used with XLSX input: {path}")
        try:
            with path.open("r", encoding="utf-8-sig", newline="") as file:
                sample = file.read(8192)
                file.seek(0)
                try:
                    dialect = csv.Sniffer().sniff(sample)
                except csv.Error:
                    dialect = csv.excel
                reader = csv.reader(file, dialect=dialect)
                first_row = next(reader, None)
                headers = validate_headers(first_row or [], path)
                rows: list[dict[str, Any]] = []
                for values in reader:
                    if not any(value not in (None, "") for value in values):
                        continue
                    rows.append(
                        {
                            header: values[index] if index < len(values) else None
                            for index, header in enumerate(headers)
                        }
                    )
                return headers, rows
        except UnicodeDecodeError as exc:
            raise PipelineError(
                f"Only UTF-8 or UTF-8-SIG CSV is supported: {path}"
            ) from exc
        except PermissionError as exc:
            raise PipelineError(f"Cannot read input file due to permissions: {path}") from exc
        except csv.Error as exc:
            raise PipelineError(f"Could not parse CSV {path}: {exc}") from exc
        except OSError as exc:
            raise PipelineError(f"Could not read input file {path}: {exc}") from exc
    if suffix == ".xlsx":
        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
            try:
                non_empty_sheets = [
                    worksheet.title
                    for worksheet in workbook.worksheets
                    if _worksheet_has_content(worksheet)
                ]
                if not non_empty_sheets:
                    raise PipelineError(f"Workbook has no non-empty worksheet: {path}")
                if sheet_name:
                    if sheet_name not in workbook.sheetnames:
                        raise PipelineError(
                            f"Worksheet '{sheet_name}' does not exist in {path.name}. "
                            f"Available worksheets: {', '.join(workbook.sheetnames)}"
                        )
                    if sheet_name not in non_empty_sheets:
                        raise PipelineError(
                            f"Worksheet '{sheet_name}' is empty in {path.name}."
                        )
                    selected_sheet = sheet_name
                elif len(non_empty_sheets) > 1:
                    raise PipelineError(
                        f"Workbook has multiple non-empty worksheets: "
                        f"{', '.join(non_empty_sheets)}. "
                        f"Select one with --sheet \"{path}=WORKSHEET\"."
                    )
                else:
                    selected_sheet = non_empty_sheets[0]

                worksheet = workbook[selected_sheet]
                values = worksheet.iter_rows(values_only=True)
                first_row = next(values, None)
                headers = validate_headers(first_row or [], path)
                rows: list[dict[str, Any]] = []
                for row in values:
                    if not any(value not in (None, "") for value in row):
                        continue
                    rows.append(
                        {
                            header: row[index] if index < len(row) else None
                            for index, header in enumerate(headers)
                        }
                    )
                return headers, rows
            finally:
                workbook.close()
        except PipelineError:
            raise
        except PermissionError as exc:
            raise PipelineError(f"Cannot read input file due to permissions: {path}") from exc
        except (
            BadZipFile,
            InvalidFileException,
            KeyError,
            ValueError,
            TypeError,
            EOFError,
            ParseError,
            OSError,
        ) as exc:
            raise PipelineError(
                f"Could not read XLSX {path}. "
                "The file may be damaged, encrypted, locked, or not a valid .xlsx file."
            ) from exc
    raise PipelineError(f"Unsupported input format: {path.suffix}. Use CSV or XLSX.")


def build_normalized_record(
    raw: dict[str, Any],
    mapping: dict[str, str],
    source: str,
    source_path: Path,
    row_number: int,
) -> NormalizedRecord:
    mapped: dict[str, Any] = {field: None for field in CANONICAL_FIELDS}
    for source_column, canonical_field in mapping.items():
        mapped[canonical_field] = raw.get(source_column)

    invalid: dict[str, str] = {}
    followers, error = parse_number(mapped["followers"])
    if error:
        invalid["followers"] = error
    average_views, error = parse_number(mapped["average_views"])
    if error:
        invalid["average_views"] = error
    engagement_rate, error = parse_number(mapped["engagement_rate"], percentage=True)
    if error:
        invalid["engagement_rate"] = error

    date_values: dict[str, datetime | None] = {}
    for field_name in (
        "latest_post_at",
        "latest_sponsored_post_at",
        "observed_at",
        "email_observed_at",
    ):
        parsed, error = parse_datetime(mapped[field_name])
        date_values[field_name] = parsed
        if error:
            invalid[field_name] = error

    email, error = normalize_email(mapped["email"])
    if error:
        invalid["email"] = error

    source_record_id = clean_text(mapped["source_record_id"]) or f"{source_path.name}:{row_number}"
    return NormalizedRecord(
        source=source,
        source_path=str(source_path),
        row_number=row_number,
        source_record_id=source_record_id,
        platform=normalize_platform(mapped["platform"]),
        platform_creator_id=clean_text(mapped["platform_creator_id"]),
        handle=normalize_handle(mapped["handle"]),
        profile_url=normalize_url(mapped["profile_url"]),
        display_name=clean_text(mapped["display_name"]),
        followers=int(followers) if followers is not None else None,
        average_views=int(average_views) if average_views is not None else None,
        engagement_rate=float(engagement_rate) if engagement_rate is not None else None,
        country=normalize_country(mapped["country"]),
        language=normalize_language(mapped["language"]),
        email=email,
        email_role=(clean_text(mapped["email_role"]) or "").lower() or None,
        latest_post_at=date_values["latest_post_at"],
        latest_sponsored_post_at=date_values["latest_sponsored_post_at"],
        observed_at=date_values["observed_at"],
        email_observed_at=date_values["email_observed_at"],
        is_estimated=parse_bool(mapped["is_estimated"]),
        raw_payload=raw,
        invalid_fields=invalid,
    )


def prepare_input_mappings(
    paths: Iterable[Path],
    *,
    sheet_names: dict[Path, str] | None = None,
    mapping_dir: Path | None = None,
    explicit_configs: list[Path] | None = None,
    interactive: bool = False,
    input_fn=input,
    output_fn=print,
    preview_fn=None,
    save_confirmed: bool = True,
) -> list[InputMappingPlan]:
    plans: list[InputMappingPlan] = []
    resolved_sheet_names = {
        path.resolve(): name for path, name in (sheet_names or {}).items()
    }
    for path in paths:
        sheet_name = resolved_sheet_names.get(path.resolve())
        headers, rows = read_tabular(path, sheet_name=sheet_name)
        try:
            plan = create_mapping_plan(
                path,
                headers,
                rows,
                sheet_name=sheet_name,
                mapping_dir=mapping_dir,
                explicit_configs=explicit_configs,
            )
            if preview_fn:
                preview_fn(plan)
            if plan.ambiguous:
                if not interactive:
                    columns = ", ".join(item.source_column for item in plan.ambiguous)
                    raise PipelineError(
                        f"{path.name} has ambiguous columns: {columns}. "
                        "Run in an interactive terminal to confirm them, or pass --mapping."
                    )
                confirm_ambiguous_columns(
                    plan,
                    input_fn=input_fn,
                    output_fn=output_fn,
                )
                plan.origin = "confirmed"
                if mapping_dir and save_confirmed:
                    plan.reused_config = save_mapping_config(plan, mapping_dir)
            else:
                validate_mapping(plan)
        except (ValueError, json.JSONDecodeError) as exc:
            raise PipelineError(str(exc)) from exc
        plans.append(plan)
    return plans


def load_records(
    paths: Iterable[Path],
    mapping_plans: list[InputMappingPlan] | None = None,
) -> tuple[list[NormalizedRecord], list[dict[str, Any]]]:
    records: list[NormalizedRecord] = []
    mapping_summaries: list[dict[str, Any]] = []
    path_list = list(paths)
    if mapping_plans is None:
        mapping_plans = prepare_input_mappings(path_list)
    plans_by_path = {plan.input_path.resolve(): plan for plan in mapping_plans}
    for path in path_list:
        plan = plans_by_path.get(path.resolve())
        if plan is None:
            raise PipelineError(f"No mapping plan was prepared for {path}.")
        headers, rows = read_tabular(path, sheet_name=plan.sheet_name)
        source = plan.source
        mapping = plan.mapping
        mapping_summaries.append(
            {
                "file": str(path),
                "source": source,
                "mapping_origin": plan.origin,
                "mapping_fingerprint": plan.fingerprint,
                "worksheet": plan.sheet_name,
                "adapter_validation_level": plan.adapter_validation_level,
                "row_count": len(rows),
                "mapped_columns": len(mapping),
                "unmapped_columns": [header for header in headers if header not in mapping],
            }
        )
        for row_number, raw in enumerate(rows, start=2):
            records.append(build_normalized_record(raw, mapping, source, path, row_number))
    if not records:
        raise PipelineError("No data rows were found in the provided inputs.")
    return records, mapping_summaries


CAMPAIGN_MARKET_ALIASES = (
    ("US", (r"美国", r"\bunited states\b", r"\busa\b", r"\bus\b")),
    ("CA", (r"加拿大", r"\bcanada\b")),
    ("BR", (r"巴西", r"\bbrazil\b")),
    ("JP", (r"日本", r"\bjapan\b")),
    ("DE", (r"德国", r"\bgermany\b")),
)

CAMPAIGN_LANGUAGE_ALIASES = (
    ("en", (r"英语", r"\benglish\b")),
    ("es", (r"西班牙语", r"\bspanish\b")),
    ("pt", (r"葡萄牙语", r"\bportuguese\b")),
    ("ja", (r"日语", r"\bjapanese\b")),
    ("de", (r"德语", r"\bgerman\b")),
)

CAMPAIGN_PLATFORM_ALIASES = (
    ("tiktok", (r"\btiktok\b", r"\btik tok\b", r"抖音海外版")),
    ("instagram", (r"\binstagram\b", r"\big\b")),
    ("youtube", (r"\byoutube\b", r"\byt\b")),
)

CAMPAIGN_TOPIC_ALIASES = (
    ("pet_lifestyle", (r"宠物日常", r"\bpet lifestyle\b")),
    ("dog_tutorial", (r"养狗教程", r"\bdog tutorial\b")),
    ("pet_tech", (r"宠物科技", r"\bpet tech\b")),
    ("smart_hardware", (r"智能硬件(?:体验)?", r"\bsmart hardware\b")),
    ("product_review", (r"产品测评", r"\bproduct reviews?\b")),
)

CAMPAIGN_QUANTITY = r"\d+(?:,\d{3})*(?:\.\d+)?\s*(?:[kKmM万])?"
CAMPAIGN_FOLLOWER_RANGE = re.compile(
    rf"(?:粉丝(?:量|范围)?|followers?|follower\s+range)"
    rf"\s*(?:为|是|[:：])?\s*({CAMPAIGN_QUANTITY})"
    rf"\s*(?:至|到|-|–|—|~|～)\s*({CAMPAIGN_QUANTITY})",
    re.IGNORECASE,
)


def _campaign_segments(raw_text: str) -> list[str]:
    segments: list[str] = []
    for line in raw_text.splitlines():
        cleaned = re.sub(r"^\s*[-*•]\s*", "", line).strip()
        if not cleaned:
            continue
        for segment in re.split(r"[。！？!?；;]+", cleaned):
            normalized = segment.strip(" \t\r\n。；;")
            if normalized:
                segments.append(normalized)
    return segments


def _matches_alias(segment: str, patterns: Iterable[str]) -> bool:
    return any(re.search(pattern, segment, re.IGNORECASE) for pattern in patterns)


def _parse_campaign_quantity(value: str) -> int:
    normalized = re.sub(r"[\s,]", "", value)
    multiplier = 1
    suffix = normalized[-1:].lower()
    if suffix == "k":
        multiplier = 1_000
        normalized = normalized[:-1]
    elif suffix == "m":
        multiplier = 1_000_000
        normalized = normalized[:-1]
    elif normalized.endswith("万"):
        multiplier = 10_000
        normalized = normalized[:-1]
    return int(round(float(normalized) * multiplier))


def parse_campaign_result(path: Path) -> CampaignParseResult:
    if not path.exists():
        raise PipelineError(f"Campaign brief does not exist: {path}")
    try:
        raw_text = path.read_text(encoding="utf-8")
    except UnicodeDecodeError as exc:
        raise PipelineError(f"Campaign brief must be UTF-8: {path}") from exc
    except PermissionError as exc:
        raise PipelineError(
            f"Cannot read Campaign brief due to permissions: {path}"
        ) from exc
    if not raw_text.strip():
        raise PipelineError(f"Campaign brief is empty: {path}")

    brand_match = re.search(
        r"(?:企业|品牌|brand)\s*[:：]?\s*([A-Za-z][A-Za-z0-9_-]+)",
        raw_text,
        re.IGNORECASE,
    )
    brand = brand_match.group(1) if brand_match else "Unspecified Brand"
    segments = _campaign_segments(raw_text)

    market_values: set[str] = set()
    language_values: set[str] = set()
    platform_values: set[str] = set()
    topic_values: set[str] = set()
    blocklist_values: set[str] = set()
    follower_ranges: list[tuple[int, int, str]] = []
    latest_values: list[tuple[int, str]] = []
    contact_values: list[tuple[bool, str]] = []
    evidence_by_field: dict[str, list[str]] = defaultdict(list)
    unrecognized: list[dict[str, str]] = []

    def note(field_name: str, evidence: str) -> None:
        if evidence not in evidence_by_field[field_name]:
            evidence_by_field[field_name].append(evidence)

    for segment in segments:
        recognized_here: set[str] = set()
        for code, patterns in CAMPAIGN_MARKET_ALIASES:
            if _matches_alias(segment, patterns):
                market_values.add(code)
                recognized_here.add("markets")
                note("markets", segment)
        for code, patterns in CAMPAIGN_LANGUAGE_ALIASES:
            if _matches_alias(segment, patterns):
                language_values.add(code)
                recognized_here.add("languages")
                note("languages", segment)
        for code, patterns in CAMPAIGN_PLATFORM_ALIASES:
            if _matches_alias(segment, patterns):
                platform_values.add(code)
                recognized_here.add("platforms")
                note("platforms", segment)
        topic_context = bool(
            re.search(
                r"内容(?:主题|方向)|topics?",
                segment,
                re.IGNORECASE,
            )
        )
        if topic_context:
            for code, patterns in CAMPAIGN_TOPIC_ALIASES:
                if _matches_alias(segment, patterns):
                    topic_values.add(code)
                    recognized_here.add("topics")
                    note("topics", segment)

        for match in CAMPAIGN_FOLLOWER_RANGE.finditer(segment):
            lower = _parse_campaign_quantity(match.group(1))
            upper = _parse_campaign_quantity(match.group(2))
            follower_ranges.append((lower, upper, segment))
            recognized_here.add("follower_range")
            note("follower_range", segment)

        latest_matches = re.findall(
            r"(?:最近|last|within)\s*(\d+)\s*(?:天|days?)",
            segment,
            re.IGNORECASE,
        )
        for value in latest_matches:
            latest_values.append((int(value), segment))
            recognized_here.add("latest_post_max_age_days")
            note("latest_post_max_age_days", segment)

        negative_contact = bool(
            re.search(
                r"(?:不需要|无需).{0,20}(?:商务)?联系"
                r"|(?:business\s+)?contact.{0,12}not\s+required",
                segment,
                re.IGNORECASE,
            )
        )
        positive_contact = bool(
            re.search(
                r"(?:需要|必须).{0,20}(?:商务)?联系"
                r"|require.{0,20}(?:business\s+)?contact"
                r"|(?:business\s+)?contact.{0,12}(?:is\s+)?required",
                segment,
                re.IGNORECASE,
            )
        )
        if negative_contact:
            contact_values.append((False, segment))
            recognized_here.add("require_contact_path")
            note("require_contact_path", segment)
        if positive_contact and not negative_contact:
            contact_values.append((True, segment))
            recognized_here.add("require_contact_path")
            note("require_contact_path", segment)

        blocklist_match = re.search(
            r"(?:排除达人|blocklist)\s*[:：]?\s*(.+)",
            segment,
            re.IGNORECASE,
        )
        if blocklist_match:
            handles = re.findall(r"@([A-Za-z0-9_.-]+)", blocklist_match.group(1))
            if handles:
                blocklist_values.update(handle.casefold() for handle in handles)
                recognized_here.add("blocklist")
                note("blocklist", segment)

        explicit_unsupported = bool(
            re.search(
                r"不要太商业化|小红书感|内容有质感|适合女性用户|女性用户"
                r"|not\s+too\s+commercial|premium\s+feel",
                segment,
                re.IGNORECASE,
            )
        )
        condition_cue = bool(
            re.search(
                r"市场|market|语言|language|平台|platform|粉丝|followers?"
                r"|最近|latest|last\s+\d+\s+days?|联系|contact|排除|blocklist"
                r"|内容|topic|不要|最好|必须|偏好|适合",
                segment,
                re.IGNORECASE,
            )
        )
        output_instruction = bool(re.search(r"^(?:请输出|输出|report\b)", segment, re.IGNORECASE))
        structured_hard_cue = bool(
            re.search(
                r"市场|market|语言|language|平台|platform|粉丝|followers?"
                r"|最近|latest|last\s+\d+\s+days?|联系|contact|排除|blocklist",
                segment,
                re.IGNORECASE,
            )
        )
        if not output_instruction and (
            explicit_unsupported or (condition_cue and not recognized_here)
        ):
            severity = (
                "blocking"
                if structured_hard_cue
                or re.search(
                    r"不要|不得|必须|排除|need|must|exclude|not\s+",
                    segment,
                    re.IGNORECASE,
                )
                else "advisory"
            )
            unrecognized.append(
                {
                    "text": segment,
                    "severity": severity,
                    "reason": "unsupported Campaign condition",
                }
            )

    conflicts: list[dict[str, Any]] = []
    unique_ranges = sorted({(lower, upper) for lower, upper, _ in follower_ranges})
    follower_min: int | None = None
    follower_max: int | None = None
    if unique_ranges:
        invalid_ranges = [item for item in unique_ranges if item[0] > item[1]]
        if len(unique_ranges) > 1 or invalid_ranges:
            conflicts.append(
                {
                    "field": "follower_range",
                    "values": [list(item) for item in unique_ranges],
                    "reason": "multiple or reversed follower ranges",
                }
            )
        else:
            follower_min, follower_max = unique_ranges[0]

    unique_latest = sorted({value for value, _ in latest_values})
    latest_post_max_age_days: int | None = None
    if len(unique_latest) > 1:
        conflicts.append(
            {
                "field": "latest_post_max_age_days",
                "values": unique_latest,
                "reason": "multiple latest-post limits",
            }
        )
    elif unique_latest:
        latest_post_max_age_days = unique_latest[0]

    unique_contact = {value for value, _ in contact_values}
    require_contact_path: bool | None = None
    if len(unique_contact) > 1:
        conflicts.append(
            {
                "field": "require_contact_path",
                "values": sorted(unique_contact),
                "reason": "contact path is both required and not required",
            }
        )
    elif unique_contact:
        require_contact_path = next(iter(unique_contact))

    markets = tuple(code for code, _ in CAMPAIGN_MARKET_ALIASES if code in market_values)
    languages = tuple(
        code for code, _ in CAMPAIGN_LANGUAGE_ALIASES if code in language_values
    )
    platforms = tuple(
        code for code, _ in CAMPAIGN_PLATFORM_ALIASES if code in platform_values
    )
    topics = tuple(code for code, _ in CAMPAIGN_TOPIC_ALIASES if code in topic_values)
    blocklist = tuple(sorted(blocklist_values))

    conditions: list[dict[str, Any]] = []

    def add_condition(
        field_name: str,
        value: Any,
        *,
        applied: bool = True,
    ) -> None:
        if not evidence_by_field.get(field_name):
            return
        conditions.append(
            {
                "field": field_name,
                "value": value,
                "evidence": list(evidence_by_field[field_name]),
                "applied": applied,
            }
        )

    add_condition("markets", list(markets))
    add_condition("languages", list(languages))
    add_condition("platforms", list(platforms))
    add_condition(
        "follower_range",
        (
            [follower_min, follower_max]
            if follower_min is not None and follower_max is not None
            else [list(item) for item in unique_ranges]
        ),
        applied=not any(item["field"] == "follower_range" for item in conflicts),
    )
    add_condition(
        "latest_post_max_age_days",
        latest_post_max_age_days if latest_post_max_age_days is not None else unique_latest,
        applied=not any(
            item["field"] == "latest_post_max_age_days" for item in conflicts
        ),
    )
    add_condition(
        "require_contact_path",
        require_contact_path if require_contact_path is not None else sorted(unique_contact),
        applied=not any(item["field"] == "require_contact_path" for item in conflicts),
    )
    add_condition("blocklist", list(blocklist))
    add_condition("topics", list(topics), applied=False)

    warnings: list[str] = []
    if unrecognized:
        warnings.append(
            "Unrecognized Campaign conditions do not participate in filtering or scoring."
        )
    if topics:
        warnings.append(
            "Topic tags are retained for review but are not applied because creator topic evidence is unavailable."
        )
    if conflicts:
        warnings.append(
            "Conflicting Campaign conditions must be resolved before a Run."
        )

    return CampaignParseResult(
        parsed_campaign=Campaign(
            brand=brand,
            markets=markets,
            languages=languages,
            platforms=platforms,
            follower_min=follower_min,
            follower_max=follower_max,
            latest_post_max_age_days=latest_post_max_age_days,
            require_contact_path=require_contact_path,
            raw_text=raw_text,
            blocklist=blocklist,
            topics=topics,
        ),
        recognized_conditions=tuple(conditions),
        unrecognized_conditions=tuple(unrecognized),
        conflicting_conditions=tuple(conflicts),
        warnings=tuple(warnings),
    )


def parse_campaign(path: Path) -> Campaign:
    """Backward-compatible access to the structured parser's Campaign value."""
    return parse_campaign_result(path).parsed_campaign


def contradictory(left: NormalizedRecord, right: NormalizedRecord) -> list[str]:
    reasons: list[str] = []
    if (
        left.platform_creator_id
        and right.platform_creator_id
        and left.platform_creator_id != right.platform_creator_id
    ):
        reasons.append("platform_creator_id differs")
    if left.profile_url and right.profile_url and left.profile_url != right.profile_url:
        reasons.append("profile_url differs")
    return reasons


def review_candidate_ids(item: dict[str, str]) -> list[str]:
    values = [item.get("candidate_a", "")] + item.get("candidate_b", "").split(",")
    return [value.strip() for value in values if value and value.strip()]


def identity_resolution(
    records: list[NormalizedRecord],
    review_decisions: list[dict[str, str]] | None = None,
) -> tuple[list[list[NormalizedRecord]], list[dict[str, str]]]:
    dsu = DisjointSet(len(records))
    reviews: dict[tuple[str, ...], dict[str, str]] = {}
    decisions = review_decisions or []
    record_index = {record.source_record_id: index for index, record in enumerate(records)}

    by_platform: dict[str, list[int]] = defaultdict(list)
    for index, record in enumerate(records):
        if record.platform:
            by_platform[record.platform].append(index)

    for indexes in by_platform.values():
        for position, left_index in enumerate(indexes):
            left = records[left_index]
            for right_index in indexes[position + 1 :]:
                right = records[right_index]
                same_platform_id = bool(
                    left.platform_creator_id
                    and left.platform_creator_id == right.platform_creator_id
                )
                same_url = bool(left.profile_url and left.profile_url == right.profile_url)
                same_handle = bool(left.handle and left.handle == right.handle)
                if not (same_platform_id or same_url or same_handle):
                    continue
                conflicts = contradictory(left, right)
                if same_platform_id:
                    dsu.union(left_index, right_index)
                    continue
                if conflicts:
                    key = tuple(sorted((left.source_record_id, right.source_record_id))) + ("identity",)
                    reviews[key] = {
                        "candidate_a": left.source_record_id,
                        "candidate_b": right.source_record_id,
                        "match_reason": "same normalized profile URL" if same_url else "same normalized handle",
                        "conflict_reason": "; ".join(conflicts),
                        "suggested_action": "keep_separate",
                        "user_decision": "",
                        "user_note": "",
                    }
                    continue
                dsu.union(left_index, right_index)

    for decision in decisions:
        if decision.get("user_decision") != "merge":
            continue
        decision_ids = review_candidate_ids(decision)
        missing_ids = [record_id for record_id in decision_ids if record_id not in record_index]
        if missing_ids:
            raise PipelineError(
                "Review input references unknown source records: "
                + ", ".join(sorted(missing_ids))
            )
        first_index = record_index[decision_ids[0]]
        for record_id in decision_ids[1:]:
            dsu.union(first_index, record_index[record_id])

    groups_by_root: dict[int, list[NormalizedRecord]] = defaultdict(list)
    for index, record in enumerate(records):
        groups_by_root[dsu.find(index)].append(record)
    groups = list(groups_by_root.values())

    by_handle: dict[str, list[NormalizedRecord]] = defaultdict(list)
    for record in records:
        if record.handle:
            by_handle[record.handle].append(record)
    for handle, candidates in by_handle.items():
        platforms = {candidate.platform for candidate in candidates if candidate.platform}
        if len(platforms) < 2:
            continue
        ids = tuple(sorted(candidate.source_record_id for candidate in candidates))
        key = ids + ("cross_platform",)
        reviews[key] = {
            "candidate_a": candidates[0].source_record_id,
            "candidate_b": ", ".join(candidate.source_record_id for candidate in candidates[1:]),
            "match_reason": f"same normalized handle across platforms: {handle}",
            "conflict_reason": "platform differs",
            "suggested_action": "keep_separate",
            "user_decision": "",
            "user_note": "",
        }

    component_by_record = {
        record.source_record_id: component_index
        for component_index, group in enumerate(groups)
        for record in group
    }
    agency_email_groups: dict[str, list[NormalizedRecord]] = defaultdict(list)
    for record in records:
        if record.email and record.email_role == "agency":
            agency_email_groups[record.email].append(record)
    for email, candidates in agency_email_groups.items():
        component_ids = {component_by_record[candidate.source_record_id] for candidate in candidates}
        if len(component_ids) < 2:
            continue
        ids = tuple(sorted(candidate.source_record_id for candidate in candidates))
        key = ids + ("agency_email",)
        reviews[key] = {
            "candidate_a": candidates[0].source_record_id,
            "candidate_b": ", ".join(candidate.source_record_id for candidate in candidates[1:]),
            "match_reason": f"shared agency email: {email}",
            "conflict_reason": "agency mailbox can represent multiple creators",
            "suggested_action": "keep_separate",
            "user_decision": "",
            "user_note": "",
        }

    decision_by_ids = {
        tuple(sorted(review_candidate_ids(item))): item for item in decisions
    }
    for item in reviews.values():
        decision = decision_by_ids.get(tuple(sorted(review_candidate_ids(item))))
        if decision:
            item["user_decision"] = decision.get("user_decision", "")
            item["user_note"] = decision.get("user_note", "")

    groups.sort(key=lambda group: min(record.source_record_id for record in group))
    return groups, sorted(reviews.values(), key=lambda item: (item["candidate_a"], item["candidate_b"]))


def observation_time(record: NormalizedRecord, field_name: str) -> datetime | None:
    if field_name == "email":
        return record.email_observed_at
    return record.observed_at


def choose_value(records: list[NormalizedRecord], field_name: str) -> tuple[Any, dict[str, Any]]:
    candidates: list[NormalizedRecord] = []
    for record in records:
        value = getattr(record, field_name)
        if value is not None and value != "":
            candidates.append(record)
    if not candidates:
        return None, {
            "source": "",
            "record_id": "",
            "observed_at": None,
            "is_estimated": False,
            "selection_reason": "No valid observation",
        }

    minimum_time = datetime.min.replace(tzinfo=timezone.utc)

    def priority(record: NormalizedRecord) -> tuple[Any, ...]:
        observed = observation_time(record, field_name)
        return (
            observed is not None,
            observed or minimum_time,
            not record.is_estimated,
            SOURCE_PRIORITY.get(record.source, 0),
            record.source_record_id,
        )

    selected_record = max(candidates, key=priority)
    observed = observation_time(selected_record, field_name)
    reason_parts = []
    reason_parts.append("known observation time" if observed else "observation time unknown")
    reason_parts.append("non-estimated" if not selected_record.is_estimated else "estimated")
    reason_parts.append(f"source={selected_record.source}")
    return getattr(selected_record, field_name), {
        "source": selected_record.source,
        "record_id": selected_record.source_record_id,
        "observed_at": observed,
        "is_estimated": selected_record.is_estimated,
        "selection_reason": "; ".join(reason_parts),
    }


def conflict_level(relative_spread: float) -> str:
    if relative_spread <= 0.05:
        return "Low"
    if relative_spread <= 0.20:
        return "Medium"
    return "High"


def create_creator_id(records: list[NormalizedRecord]) -> str:
    identity = "|".join(sorted(record.source_record_id for record in records))
    return f"creator_{hashlib.sha1(identity.encode('utf-8')).hexdigest()[:12]}"


def infer_merge_basis(records: list[NormalizedRecord]) -> str:
    if len(records) == 1:
        return "single source record"
    platform_ids = {record.platform_creator_id for record in records if record.platform_creator_id}
    urls = {record.profile_url for record in records if record.profile_url}
    handles = {record.handle for record in records if record.handle}
    if len(platform_ids) == 1:
        return "same platform + platform_creator_id"
    if len(urls) == 1:
        return "same platform + normalized profile_url"
    if len(handles) == 1:
        return "same platform + normalized handle"
    return "deterministic transitive identity keys"


def aggregate_group(
    records: list[NormalizedRecord], reference_time: datetime
) -> CreatorAggregate:
    creator_id = create_creator_id(records)
    selected: dict[str, Any] = {}
    selected_meta: dict[str, dict[str, Any]] = {}
    fields = (
        "platform",
        "platform_creator_id",
        "handle",
        "profile_url",
        "display_name",
        "followers",
        "average_views",
        "engagement_rate",
        "country",
        "language",
        "email",
        "email_role",
        "latest_post_at",
        "latest_sponsored_post_at",
    )
    for field_name in fields:
        selected[field_name], selected_meta[field_name] = choose_value(records, field_name)

    conflicts: list[Conflict] = []
    numeric_fields = ("followers", "average_views", "engagement_rate")
    for field_name in numeric_fields:
        observations = [
            (record.source_record_id, record.source, getattr(record, field_name))
            for record in records
            if getattr(record, field_name) is not None
        ]
        values = [float(item[2]) for item in observations]
        if len(values) < 2 or max(values) == min(values):
            continue
        spread = (max(values) - min(values)) / max(max(values), 1.0)
        level = conflict_level(spread)
        conflicts.append(
            Conflict(
                creator_id=creator_id,
                field=field_name,
                level=level,
                values="; ".join(f"{record_id}={value}" for record_id, _, value in observations),
                sources=", ".join(sorted({source for _, source, _ in observations})),
                reason=f"relative spread {spread:.1%}",
            )
        )

    for field_name in ("country", "language", "email"):
        observations = [
            (record.source_record_id, record.source, getattr(record, field_name))
            for record in records
            if getattr(record, field_name)
        ]
        unique_values = {item[2] for item in observations}
        if len(unique_values) < 2:
            continue
        conflicts.append(
            Conflict(
                creator_id=creator_id,
                field=field_name,
                level="High",
                values="; ".join(f"{record_id}={value}" for record_id, _, value in observations),
                sources=", ".join(sorted({source for _, source, _ in observations})),
                reason="normalized categorical values disagree",
            )
        )

    missing_fields = [field_name for field_name in CRITICAL_FIELDS if selected.get(field_name) is None]
    stale_items: list[StaleItem] = []
    for field_name, ttl_days in TTL_DAYS.items():
        value = selected.get(field_name)
        if value is None:
            continue
        meta = selected_meta[field_name]
        observed = meta["observed_at"]
        if observed is None:
            stale_items.append(
                StaleItem(
                    creator_id,
                    field_name,
                    value,
                    meta["source"],
                    None,
                    ttl_days,
                    "Observed Time Unknown",
                )
            )
        elif (reference_time - observed).days > ttl_days:
            stale_items.append(
                StaleItem(
                    creator_id,
                    field_name,
                    value,
                    meta["source"],
                    observed,
                    ttl_days,
                    "Stale",
                )
            )

    invalid_fields = sorted(
        {
            f"{record.source_record_id}:{field_name}={raw_value}"
            for record in records
            for field_name, raw_value in record.invalid_fields.items()
        }
    )
    return CreatorAggregate(
        creator_id=creator_id,
        records=records,
        selected=selected,
        selected_meta=selected_meta,
        conflicts=conflicts,
        stale_items=stale_items,
        missing_fields=missing_fields,
        invalid_fields=invalid_fields,
        merge_basis=infer_merge_basis(records),
    )


def weighted_available(features: list[tuple[float, float, bool]]) -> float:
    available = [(score, weight) for score, weight, is_available in features if is_available]
    denominator = sum(weight for _, weight in available)
    if denominator == 0:
        return 0.0
    return sum(score * weight for score, weight in available) / denominator


def score_creator(
    creator: CreatorAggregate,
    campaign: Campaign,
    reference_time: datetime,
) -> None:
    selected = creator.selected
    country = selected.get("country")
    language = selected.get("language")
    platform = selected.get("platform")
    handle = selected.get("handle")
    followers = selected.get("followers")
    latest_post_at = selected.get("latest_post_at")
    email = selected.get("email")
    email_role = selected.get("email_role")

    market_match = not campaign.markets or country in campaign.markets
    language_match = not campaign.languages or language in campaign.languages
    platform_match = not campaign.platforms or platform in campaign.platforms
    blocklisted = bool(handle and handle.casefold() in set(campaign.blocklist))
    follower_match = True
    if followers is None:
        follower_match = False
    if campaign.follower_min is not None and followers is not None and followers < campaign.follower_min:
        follower_match = False
    if campaign.follower_max is not None and followers is not None and followers > campaign.follower_max:
        follower_match = False

    latest_age_days: int | None = None
    if latest_post_at:
        latest_age_days = max(0, (reference_time - latest_post_at).days)
    active_match = bool(
        latest_age_days is not None
        and (
            campaign.latest_post_max_age_days is None
            or latest_age_days <= campaign.latest_post_max_age_days
        )
    )

    creator.exclusion_reasons = []
    if blocklisted:
        creator.exclusion_reasons.append("EXCLUDED_BLOCKLIST")
    if campaign.platforms and not platform_match:
        creator.exclusion_reasons.append("EXCLUDED_PLATFORM")
    if campaign.markets and not market_match:
        creator.exclusion_reasons.append("EXCLUDED_MARKET")
    if campaign.languages and not language_match:
        creator.exclusion_reasons.append("EXCLUDED_LANGUAGE")
    if not follower_match:
        creator.exclusion_reasons.append("EXCLUDED_FOLLOWER_RANGE")
    if campaign.latest_post_max_age_days is not None and not active_match:
        creator.exclusion_reasons.append("EXCLUDED_INACTIVE")
    if campaign.require_contact_path and not email:
        creator.exclusion_reasons.append("EXCLUDED_NO_CONTACT")

    brand_fit = weighted_available(
        [
            (100.0 if market_match else 0.0, 0.15, country is not None),
            (100.0 if language_match else 0.0, 0.15, language is not None),
            (0.0, 0.35, False),
            (0.0, 0.20, False),
            (0.0, 0.15, False),
        ]
    )
    commercial_readiness = weighted_available(
        [
            (100.0 if active_match else 0.0, 0.30, latest_post_at is not None),
            (100.0 if email else 0.0, 0.30, True),
            (100.0 if selected.get("latest_sponsored_post_at") else 0.0, 0.25, selected.get("latest_sponsored_post_at") is not None),
            (0.0, 0.15, False),
        ]
    )
    email_source_score = 100.0 if email_role == "creator" else 75.0 if email_role == "agency" else 50.0
    email_observed = creator.selected_meta.get("email", {}).get("observed_at")
    email_fresh = bool(email_observed and (reference_time - email_observed).days <= TTL_DAYS["email"])
    contactability = weighted_available(
        [
            (email_source_score if email else 0.0, 0.45, True),
            (100.0 if email_role in {"creator", "agency"} else 50.0, 0.20, email is not None),
            (100.0 if email_fresh else 0.0, 0.20, email_observed is not None),
            (0.0, 0.15, False),
        ]
    )

    covered_fields = sum(selected.get(field_name) is not None for field_name in CRITICAL_FIELDS)
    coverage_score = 100.0 * covered_fields / len(CRITICAL_FIELDS)
    freshness_status_fields = [
        field_name for field_name in TTL_DAYS if selected.get(field_name) is not None
    ]
    stale_or_unknown = {item.field for item in creator.stale_items}
    freshness_score = (
        100.0
        * sum(field_name not in stale_or_unknown for field_name in freshness_status_fields)
        / len(freshness_status_fields)
        if freshness_status_fields
        else 0.0
    )
    penalties = {"Low": 5, "Medium": 18, "High": 35}
    consistency_score = max(
        0.0,
        100.0 - sum(penalties.get(conflict.level, 0) for conflict in creator.conflicts),
    )
    source_count = len(creator.source_names)
    source_diversity = 40.0 if source_count == 1 else 80.0 if source_count == 2 else 100.0
    data_confidence = (
        0.30 * coverage_score
        + 0.30 * freshness_score
        + 0.25 * consistency_score
        + 0.15 * source_diversity
    )

    creator.dimensions = {
        "brand_fit": round(brand_fit, 2),
        "commercial_readiness": round(commercial_readiness, 2),
        "contactability": round(contactability, 2),
        "data_confidence": round(data_confidence, 2),
    }
    creator.reference_score = round(
        0.35 * brand_fit
        + 0.25 * commercial_readiness
        + 0.25 * contactability
        + 0.15 * data_confidence,
        2,
    )

    creator.warnings = []
    creator.warnings.extend(
        f"{conflict.level} conflict: {conflict.field}" for conflict in creator.conflicts
    )
    creator.warnings.extend(f"Missing: {field_name}" for field_name in creator.missing_fields)
    creator.warnings.extend(
        f"{item.status}: {item.field}" for item in creator.stale_items
    )
    creator.warnings.extend(f"Invalid: {item}" for item in creator.invalid_fields)
    if creator.review_required:
        creator.warnings.append("Identity review required")

    creator.refresh_recommendation = sorted(
        {
            item.field for item in creator.stale_items
        }
        | {
            conflict.field
            for conflict in creator.conflicts
            if conflict.level in {"Medium", "High"}
        }
        | set(creator.missing_fields)
    )

    material_conflict = any(
        conflict.level in {"Medium", "High"}
        and conflict.field in {"followers", "country", "language", "email"}
        for conflict in creator.conflicts
    )
    material_stale = any(item.field == "email" for item in creator.stale_items)

    if creator.exclusion_reasons:
        creator.action_level = "Excluded"
    elif creator.reference_score >= 75 and not (
        material_conflict or material_stale or creator.review_required
    ):
        creator.action_level = "Priority"
    elif creator.reference_score >= 60:
        creator.action_level = "Verify"
    else:
        creator.action_level = "Hold"

    reasons: list[str] = []
    if market_match and country:
        reasons.append(f"目标市场匹配：{country}")
    if language_match and language:
        reasons.append(f"内容语言匹配：{language}")
    if active_match and latest_age_days is not None:
        reasons.append(f"近期活跃：最近发帖距今 {latest_age_days} 天")
    if email:
        reasons.append(f"存在商务联系路径：{email_role or 'role unknown'}")
    if source_count >= 2:
        reasons.append(f"由 {source_count} 个来源交叉支持")
    if not reasons:
        reasons.append("当前证据不足，保留供人工复核")
    if len(reasons) == 1:
        reasons.append(f"数据置信度参考分 {data_confidence:.1f}")
    creator.why_contact = reasons[:4]


def mark_review_creators(
    creators: list[CreatorAggregate], review_items: list[dict[str, str]]
) -> None:
    record_to_creator = {
        record.source_record_id: creator
        for creator in creators
        for record in creator.records
    }
    for item in review_items:
        if item.get("user_decision") in {"merge", "keep_separate"}:
            continue
        record_ids = review_candidate_ids(item)
        for record_id in record_ids:
            creator = record_to_creator.get(record_id)
            if creator:
                creator.review_required = True


def make_run_id(input_paths: list[Path], brief_path: Path) -> str:
    digest = hashlib.sha256()
    for path in [*input_paths, brief_path]:
        digest.update(path.name.encode("utf-8"))
        digest.update(path.read_bytes())
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    return f"run_{timestamp}_{digest.hexdigest()[:8]}"


def ensure_unique_run_dir(output_base: Path, run_id: str) -> tuple[str, Path]:
    candidate = output_base / run_id
    counter = 2
    while candidate.exists():
        candidate = output_base / f"{run_id}_{counter}"
        counter += 1
    try:
        candidate.mkdir(parents=True, exist_ok=False)
    except PermissionError as exc:
        raise OutputError(f"Cannot create output directory: {candidate}") from exc
    except OSError as exc:
        raise OutputError(f"Cannot create output directory {candidate}: {exc}") from exc
    return candidate.name, candidate


def file_hash(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as file:
        for chunk in iter(lambda: file.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def validate_campaign_for_run(
    result: CampaignParseResult,
    *,
    confirmed: bool,
) -> None:
    if result.conflicting_conditions:
        fields = ", ".join(
            str(item.get("field", "unknown"))
            for item in result.conflicting_conditions
        )
        raise PipelineError(
            f"Resolve conflicting Campaign conditions before running: {fields}."
        )
    if result.blocking_unrecognized_conditions and not confirmed:
        texts = "; ".join(
            item["text"] for item in result.blocking_unrecognized_conditions
        )
        raise PipelineError(
            f"Found unrecognized Campaign condition(s): {texts}. "
            "Review them with campaign-preview, then pass a confirmed "
            "--campaign-config if you accept that they will not be applied."
        )


def write_campaign_config(
    path: Path,
    *,
    brief_path: Path,
    result: CampaignParseResult,
    confirmed: bool,
    confirmation_source: str,
) -> None:
    if confirmed and result.conflicting_conditions:
        raise PipelineError(
            "Cannot confirm a Campaign config with conflicting Campaign conditions."
        )
    payload = {
        "schema_version": 1,
        "confirmed": confirmed,
        "confirmation_source": confirmation_source,
        "source_brief": {
            "path": str(brief_path.resolve()),
            "sha256": file_hash(brief_path),
        },
        **result.to_dict(),
    }
    try:
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(
            json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )
    except PermissionError as exc:
        raise OutputError(f"Cannot write Campaign config: {path}") from exc
    except OSError as exc:
        raise OutputError(f"Cannot write Campaign config {path}: {exc}") from exc


def load_campaign_config(
    path: Path,
    *,
    brief_path: Path,
) -> CampaignParseResult:
    if not path.exists():
        raise PipelineError(f"Campaign config does not exist: {path}")
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (UnicodeDecodeError, json.JSONDecodeError) as exc:
        raise PipelineError(f"Campaign config is not valid UTF-8 JSON: {path}") from exc
    if payload.get("schema_version") != 1:
        raise PipelineError(f"Unsupported Campaign config schema: {path}")
    if payload.get("confirmed") is not True:
        raise PipelineError(f"Campaign config is not confirmed: {path}")
    expected_hash = payload.get("source_brief", {}).get("sha256")
    if expected_hash != file_hash(brief_path):
        raise PipelineError(
            "Campaign config does not match the supplied Campaign brief."
        )
    raw_text = brief_path.read_text(encoding="utf-8")
    result = CampaignParseResult.from_dict(payload, raw_text=raw_text)
    validate_campaign_for_run(result, confirmed=True)
    return result


def sort_creators(creators: list[CreatorAggregate]) -> None:
    action_order = {"Priority": 0, "Verify": 1, "Hold": 2, "Excluded": 3}
    creators.sort(
        key=lambda creator: (
            action_order[creator.action_level],
            -creator.reference_score,
            creator.creator_id,
        )
    )


def manifest_mapping_plans(plans: list[InputMappingPlan]) -> list[dict[str, Any]]:
    return [
        {
            "input_path": str(plan.input_path.resolve()),
            "fingerprint": plan.fingerprint,
            "source": plan.source,
            "origin": plan.origin,
            "sheet_name": plan.sheet_name,
            "adapter_version": plan.adapter_version,
            "adapter_validation_level": plan.adapter_validation_level,
            "input_columns": list(plan.input_columns),
            "mapping": plan.mapping,
        }
        for plan in plans
    ]


def mapping_plans_from_manifest(payload: dict[str, Any]) -> list[InputMappingPlan]:
    plans = []
    for item in payload.get("mapping_plans", []):
        plans.append(
            InputMappingPlan(
                input_path=Path(item["input_path"]),
                fingerprint=item["fingerprint"],
                source=item["source"],
                origin=item.get("origin", "manifest"),
                mapping=dict(item["mapping"]),
                suggestions=[],
                sheet_name=item.get("sheet_name"),
                adapter_version=item.get("adapter_version"),
                adapter_validation_level=item.get("adapter_validation_level"),
                input_columns=list(item.get("input_columns", item["mapping"])),
            )
        )
    return plans


def write_run_manifest(
    path: Path,
    *,
    run_id: str,
    input_paths: list[Path],
    brief_path: Path,
    mapping_plans: list[InputMappingPlan],
    campaign_parse_result: CampaignParseResult,
    campaign_confirmation_source: str,
    reference_time: datetime,
    top: int,
    review_revision: int = 0,
    review_decisions: list[dict[str, str]] | None = None,
    created_at: str | None = None,
    run_type: str = "standard",
    data_classification: str = "user_provided",
) -> None:
    now = datetime.now(timezone.utc).isoformat()
    payload = {
        "schema_version": 1,
        "run_id": run_id,
        "run_type": run_type,
        "data_classification": data_classification,
        "input_files": [
            {"path": str(path.resolve()), "sha256": file_hash(path)}
            for path in input_paths
        ],
        "brief": {
            "path": str(brief_path.resolve()),
            "sha256": file_hash(brief_path),
        },
        "campaign_parse": campaign_parse_result.to_dict(),
        "campaign_confirmation_source": campaign_confirmation_source,
        "mapping_plans": manifest_mapping_plans(mapping_plans),
        "reference_time": reference_time.isoformat(),
        "top": top,
        "review_revision": review_revision,
        "review_decisions": review_decisions or [],
        "created_at": created_at or now,
        "updated_at": now,
    }
    path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )


def load_run_manifest(run_dir: Path) -> dict[str, Any]:
    path = run_dir / "manifest.json"
    if not path.exists():
        raise PipelineError(
            f"Run {run_dir.name} has no manifest.json. "
            "Review requires a Run created by Milestone 3 or later."
        )
    payload = json.loads(path.read_text(encoding="utf-8"))
    if payload.get("schema_version") != 1:
        raise PipelineError(f"Unsupported manifest schema in {path}.")
    return payload


def verify_manifest_inputs(payload: dict[str, Any]) -> tuple[list[Path], Path]:
    input_paths = [Path(item["path"]) for item in payload.get("input_files", [])]
    brief_path = Path(payload.get("brief", {}).get("path", ""))
    for path, item in zip(input_paths, payload.get("input_files", []), strict=True):
        if not path.exists():
            raise PipelineError(f"Original input is missing: {path}")
        if file_hash(path) != item["sha256"]:
            raise PipelineError(
                f"Original input changed after the Run was created: {path}. "
                "Create a new Run instead of reviewing changed data."
            )
    if not brief_path.exists():
        raise PipelineError(f"Original Campaign brief is missing: {brief_path}")
    if file_hash(brief_path) != payload["brief"]["sha256"]:
        raise PipelineError(
            "Campaign brief changed after the Run was created. Create a new Run."
        )
    return input_paths, brief_path


def read_review_decisions(
    path: Path,
    *,
    sheet_name: str | None = None,
) -> list[dict[str, str]]:
    headers, rows = read_tabular(path, sheet_name=sheet_name)
    required = {
        "candidate_a",
        "candidate_b",
        "user_decision",
        "user_note",
    }
    missing = required - set(headers)
    if missing:
        raise PipelineError(
            "Review workbook is missing columns: " + ", ".join(sorted(missing))
        )
    decisions: list[dict[str, str]] = []
    seen: set[tuple[str, ...]] = set()
    for row_number, row in enumerate(rows, start=2):
        decision = (clean_text(row.get("user_decision")) or "").lower()
        if decision not in {"", "merge", "keep_separate", "unsure"}:
            raise PipelineError(
                f"Invalid user_decision at row {row_number}: {decision}. "
                "Use merge, keep_separate, unsure, or blank."
            )
        item = {
            "candidate_a": clean_text(row.get("candidate_a")) or "",
            "candidate_b": clean_text(row.get("candidate_b")) or "",
            "match_reason": clean_text(row.get("match_reason")) or "",
            "conflict_reason": clean_text(row.get("conflict_reason")) or "",
            "suggested_action": clean_text(row.get("suggested_action")) or "",
            "user_decision": decision,
            "user_note": clean_text(row.get("user_note")) or "",
        }
        ids = tuple(sorted(review_candidate_ids(item)))
        if len(ids) < 2:
            raise PipelineError(f"Review row {row_number} must reference two candidates.")
        if ids in seen:
            raise PipelineError(f"Duplicate review decision at row {row_number}.")
        seen.add(ids)
        decisions.append(item)
    return decisions


def run_pipeline(
    input_paths: list[Path],
    brief_path: Path,
    output_base: Path,
    top: int = 20,
    mapping_plans: list[InputMappingPlan] | None = None,
    campaign_result: CampaignParseResult | None = None,
    campaign_confirmed: bool = False,
    campaign_confirmation_source: str = "auto-validated",
    reference_time: datetime | None = None,
    run_type: str = "standard",
    data_classification: str = "user_provided",
) -> PipelineResult:
    if not input_paths:
        raise PipelineError("At least one --input file is required.")
    campaign_result = campaign_result or parse_campaign_result(brief_path)
    validate_campaign_for_run(campaign_result, confirmed=campaign_confirmed)
    campaign = campaign_result.parsed_campaign
    if mapping_plans is None:
        mapping_plans = prepare_input_mappings(input_paths)
    records, mapping_summaries = load_records(input_paths, mapping_plans=mapping_plans)
    groups, review_items = identity_resolution(records)
    reference_time = reference_time or datetime.now(timezone.utc).replace(
        hour=0,
        minute=0,
        second=0,
        microsecond=0,
    )
    creators = [aggregate_group(group, reference_time) for group in groups]
    mark_review_creators(creators, review_items)
    for creator in creators:
        score_creator(creator, campaign, reference_time)

    sort_creators(creators)

    run_id, run_dir = ensure_unique_run_dir(output_base, make_run_id(input_paths, brief_path))
    audit_path = run_dir / "data_audit.xlsx"
    shortlist_path = run_dir / "creator_shortlist.xlsx"
    review_path = run_dir / "merge_review.xlsx"
    feedback_template_path = run_dir / "feedback_template.xlsx"
    report_path = run_dir / "report.md"
    campaign_config_path = run_dir / "campaign.json"
    manifest_path = run_dir / "manifest.json"

    write_data_audit(
        audit_path,
        records=records,
        creators=creators,
        review_items=review_items,
        mapping_summaries=mapping_summaries,
        reference_time=reference_time,
    )
    write_creator_shortlist(shortlist_path, creators)
    write_merge_review(review_path, review_items)
    write_feedback_template(feedback_template_path, creators)
    write_report(
        report_path,
        run_id=run_id,
        campaign=campaign,
        campaign_parse_result=campaign_result,
        records=records,
        creators=creators,
        review_items=review_items,
        mapping_summaries=mapping_summaries,
        reference_time=reference_time,
        top=top,
        run_type=run_type,
        data_classification=data_classification,
    )
    write_campaign_config(
        campaign_config_path,
        brief_path=brief_path,
        result=campaign_result,
        confirmed=True,
        confirmation_source=campaign_confirmation_source,
    )
    write_run_manifest(
        manifest_path,
        run_id=run_id,
        input_paths=input_paths,
        brief_path=brief_path,
        mapping_plans=mapping_plans,
        campaign_parse_result=campaign_result,
        campaign_confirmation_source=campaign_confirmation_source,
        reference_time=reference_time,
        top=top,
        run_type=run_type,
        data_classification=data_classification,
    )

    return PipelineResult(
        run_id=run_id,
        run_dir=run_dir,
        raw_record_count=len(records),
        creator_count=len(creators),
        review_count=len(review_items),
        output_files=(
            audit_path,
            shortlist_path,
            review_path,
            feedback_template_path,
            report_path,
        ),
    )


def review_run(
    *,
    run_id: str,
    review_input: Path,
    runs_dir: Path = Path("runs"),
    review_sheet_name: str | None = None,
) -> PipelineResult:
    run_dir = runs_dir / run_id
    if not run_dir.is_dir():
        raise PipelineError(f"Run does not exist: {run_id}")
    manifest = load_run_manifest(run_dir)
    input_paths, brief_path = verify_manifest_inputs(manifest)
    mapping_plans = mapping_plans_from_manifest(manifest)
    review_decisions = read_review_decisions(
        review_input,
        sheet_name=review_sheet_name,
    )
    raw_brief = brief_path.read_text(encoding="utf-8")
    if isinstance(manifest.get("campaign_parse"), dict):
        campaign_result = CampaignParseResult.from_dict(
            manifest["campaign_parse"],
            raw_text=raw_brief,
        )
    else:
        campaign_result = parse_campaign_result(brief_path)
    validate_campaign_for_run(campaign_result, confirmed=True)
    campaign = campaign_result.parsed_campaign
    campaign_confirmation_source = str(
        manifest.get("campaign_confirmation_source", "legacy-run")
    )
    records, mapping_summaries = load_records(
        input_paths,
        mapping_plans=mapping_plans,
    )
    groups, review_items = identity_resolution(
        records,
        review_decisions=review_decisions,
    )
    reference_time = datetime.fromisoformat(manifest["reference_time"])
    creators = [aggregate_group(group, reference_time) for group in groups]
    manual_merge_sets = [
        set(review_candidate_ids(item))
        for item in review_decisions
        if item.get("user_decision") == "merge"
    ]
    for creator in creators:
        creator_records = set(creator.source_record_ids)
        if any(decision_ids.issubset(creator_records) for decision_ids in manual_merge_sets):
            creator.merge_basis = "manual review merge"
    mark_review_creators(creators, review_items)
    for creator in creators:
        score_creator(creator, campaign, reference_time)
    sort_creators(creators)

    audit_path = run_dir / "data_audit.xlsx"
    shortlist_path = run_dir / "creator_shortlist.xlsx"
    review_path = run_dir / "merge_review.xlsx"
    feedback_template_path = run_dir / "feedback_template.xlsx"
    report_path = run_dir / "report.md"
    write_data_audit(
        audit_path,
        records=records,
        creators=creators,
        review_items=review_items,
        mapping_summaries=mapping_summaries,
        reference_time=reference_time,
    )
    write_creator_shortlist(shortlist_path, creators)
    write_merge_review(review_path, review_items)
    write_feedback_template(feedback_template_path, creators)
    write_report(
        report_path,
        run_id=run_id,
        campaign=campaign,
        campaign_parse_result=campaign_result,
        records=records,
        creators=creators,
        review_items=review_items,
        mapping_summaries=mapping_summaries,
        reference_time=reference_time,
        top=int(manifest.get("top", 20)),
        run_type=str(manifest.get("run_type", "standard")),
        data_classification=str(
            manifest.get("data_classification", "user_provided")
        ),
    )
    write_campaign_config(
        run_dir / "campaign.json",
        brief_path=brief_path,
        result=campaign_result,
        confirmed=True,
        confirmation_source=campaign_confirmation_source,
    )
    write_run_manifest(
        run_dir / "manifest.json",
        run_id=run_id,
        input_paths=input_paths,
        brief_path=brief_path,
        mapping_plans=mapping_plans,
        campaign_parse_result=campaign_result,
        campaign_confirmation_source=campaign_confirmation_source,
        reference_time=reference_time,
        top=int(manifest.get("top", 20)),
        review_revision=int(manifest.get("review_revision", 0)) + 1,
        review_decisions=review_decisions,
        created_at=manifest.get("created_at"),
        run_type=str(manifest.get("run_type", "standard")),
        data_classification=str(
            manifest.get("data_classification", "user_provided")
        ),
    )
    return PipelineResult(
        run_id=run_id,
        run_dir=run_dir,
        raw_record_count=len(records),
        creator_count=len(creators),
        review_count=sum(
            item.get("user_decision") not in {"merge", "keep_separate"}
            for item in review_items
        ),
        output_files=(
            audit_path,
            shortlist_path,
            review_path,
            feedback_template_path,
            report_path,
        ),
    )


def serialize_debug_summary(creators: list[CreatorAggregate]) -> str:
    """Compact deterministic representation used by tests, not a persisted product artifact."""
    payload = [
        {
            "records": creator.source_record_ids,
            "action": creator.action_level,
            "score": creator.reference_score,
            "conflicts": sorted((item.field, item.level) for item in creator.conflicts),
        }
        for creator in creators
    ]
    return json.dumps(payload, ensure_ascii=False, sort_keys=True)
