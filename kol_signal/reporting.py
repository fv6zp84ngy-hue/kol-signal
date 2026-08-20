from __future__ import annotations

from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import TYPE_CHECKING, Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

if TYPE_CHECKING:
    from .core import (
        Campaign,
        CampaignParseResult,
        CreatorAggregate,
        NormalizedRecord,
    )


DARK_BLUE = "1F4E78"
MID_BLUE = "5B9BD5"
LIGHT_BLUE = "DDEBF7"
LIGHT_GRAY = "F3F6F8"
WHITE = "FFFFFF"
TEXT = "1F2937"
GREEN = "E2F0D9"
YELLOW = "FFF2CC"
ORANGE = "FCE4D6"
RED = "F4CCCC"
PURPLE = "E4DFEC"
THIN_GRAY = Side(style="thin", color="D9E2F3")


EXCEL_FORMULA_PREFIXES = ("=", "+", "-", "@", "\t", "\r")


def safe_excel_value(value: Any) -> Any:
    """Return a value that openpyxl cannot interpret as a user-supplied formula."""
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    if isinstance(value, str) and value.startswith(EXCEL_FORMULA_PREFIXES):
        return f"'{value}"
    return value


def append_safe_row(worksheet, values: Iterable[Any]) -> None:
    worksheet.append([safe_excel_value(value) for value in values])


def stringify(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M")
    return str(value)


def sanitize_table_name(name: str) -> str:
    return "".join(character for character in name if character.isalnum() or character == "_")[:200]


def style_header(worksheet, row: int, start_column: int, end_column: int) -> None:
    for cell in worksheet.iter_cols(
        min_col=start_column,
        max_col=end_column,
        min_row=row,
        max_row=row,
    ):
        target = cell[0]
        target.fill = PatternFill("solid", fgColor=DARK_BLUE)
        target.font = Font(name="Aptos", size=10, bold=True, color=WHITE)
        target.alignment = Alignment(vertical="center", wrap_text=True)
        target.border = Border(bottom=Side(style="medium", color=DARK_BLUE))
    worksheet.row_dimensions[row].height = 30


def set_column_widths(worksheet, headers: list[str], max_row: int) -> None:
    for index, header in enumerate(headers, start=1):
        normalized = header.lower()
        if normalized in {"why_contact", "data_warnings", "refresh_recommendation", "exclusion_reasons"}:
            width = 42
        elif normalized in {"profile_url", "values", "selection_reason", "conflict_reason", "user_note"}:
            width = 38
        elif "email" in normalized:
            width = 28
        elif "record" in normalized or normalized in {"creator_id", "candidate_a", "candidate_b"}:
            width = 25
        elif "date" in normalized or "observed" in normalized:
            width = 20
        elif normalized in {"sources_used", "source_record_ids", "unmapped_columns"}:
            width = 30
        else:
            width = max(12, min(22, len(header) + 3))
        worksheet.column_dimensions[get_column_letter(index)].width = width

    for row in worksheet.iter_rows(min_row=2, max_row=max_row):
        for cell in row:
            cell.font = Font(name="Aptos", size=10, color=TEXT)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=THIN_GRAY)


def add_table(worksheet, name: str, start_row: int, end_row: int, end_column: int) -> None:
    if end_row < start_row:
        return
    ref = f"A{start_row}:{get_column_letter(end_column)}{end_row}"
    table = Table(displayName=sanitize_table_name(name), ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)


def write_rows(
    worksheet,
    headers: list[str],
    rows: Iterable[Iterable[Any]],
    *,
    table_name: str,
    freeze: str = "A2",
) -> int:
    worksheet.sheet_view.showGridLines = False
    append_safe_row(worksheet, headers)
    row_count = 1
    for row in rows:
        append_safe_row(worksheet, row)
        row_count += 1
    style_header(worksheet, 1, 1, len(headers))
    set_column_widths(worksheet, headers, row_count)
    worksheet.freeze_panes = freeze
    add_table(worksheet, table_name, 1, row_count, len(headers))
    return row_count


def creator_selected_value(creator: CreatorAggregate, field_name: str) -> Any:
    return creator.selected.get(field_name)


def write_data_audit(
    path: Path,
    *,
    records: list[NormalizedRecord],
    creators: list[CreatorAggregate],
    review_items: list[dict[str, str]],
    mapping_summaries: list[dict[str, Any]],
    reference_time: datetime,
) -> None:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    summary.sheet_view.showGridLines = False
    summary.merge_cells("A1:C1")
    summary["A1"] = "Creator List Data Audit"
    summary["A1"].fill = PatternFill("solid", fgColor=DARK_BLUE)
    summary["A1"].font = Font(name="Aptos Display", size=16, bold=True, color=WHITE)
    summary["A1"].alignment = Alignment(vertical="center")
    summary.row_dimensions[1].height = 34
    append_safe_row(summary, [])
    append_safe_row(summary, ["Metric", "Value", "Notes"])

    conflict_count = sum(len(creator.conflicts) for creator in creators)
    high_conflict_count = sum(
        conflict.level == "High" for creator in creators for conflict in creator.conflicts
    )
    stale_count = sum(len(creator.stale_items) for creator in creators)
    invalid_count = sum(len(record.invalid_fields) for record in records)
    email_count = sum(bool(creator.selected.get("email")) for creator in creators)
    duplicate_groups = sum(len(creator.records) > 1 for creator in creators)
    unresolved_reviews = sum(
        item.get("user_decision") not in {"merge", "keep_separate"}
        for item in review_items
    )
    resolved_reviews = len(review_items) - unresolved_reviews
    summary_rows = [
        ("Reference time", reference_time, "Freshness is evaluated against this run timestamp."),
        ("Input files", len(mapping_summaries), "Mappings came from native, confirmed, or reused rules."),
        ("Raw records", len(records), "Rows read from all CSV/XLSX files."),
        ("Canonical creators", len(creators), "Creators after deterministic same-platform merging."),
        ("Records removed by dedup", len(records) - len(creators), "Raw records minus canonical creators."),
        ("Merged creator groups", duplicate_groups, "Creator identities supported by at least two source rows."),
        ("Pending review items", unresolved_reviews, "Identity signals still awaiting a user decision."),
        ("Resolved review items", resolved_reviews, "Rows confirmed as merge or keep_separate."),
        ("Email coverage", email_count / len(creators) if creators else 0, "Canonical creators with a selected email."),
        ("Field conflicts", conflict_count, "Low, Medium, and High field-level disagreements."),
        ("High conflicts", high_conflict_count, "Conflicts likely to affect identity, filtering, or contact."),
        ("Stale/unknown observations", stale_count, "TTL failures plus missing observation times."),
        ("Invalid source values", invalid_count, "Invalid values remain visible and are never converted to zero."),
    ]
    for row in summary_rows:
        append_safe_row(summary, row)
    style_header(summary, 3, 1, 3)
    summary.column_dimensions["A"].width = 30
    summary.column_dimensions["B"].width = 22
    summary.column_dimensions["C"].width = 64
    summary.freeze_panes = "A4"
    for row in summary.iter_rows(min_row=4, max_row=summary.max_row):
        for cell in row:
            cell.font = Font(name="Aptos", size=10, color=TEXT)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=THIN_GRAY)
    summary["B12"].number_format = "0.0%"
    add_table(summary, "AuditSummary", 3, summary.max_row, 3)

    source_sheet = workbook.create_sheet("Source Comparison")
    source_headers = [
        "source",
        "raw_records",
        "canonical_creators_present",
        "source_unique_creators",
        "overlap_creators",
        "email_records",
        "email_record_coverage",
        "records_with_invalid_values",
    ]
    source_rows = []
    source_names = sorted({record.source for record in records})
    for source in source_names:
        source_records = [record for record in records if record.source == source]
        source_creators = [creator for creator in creators if source in creator.source_names]
        source_unique = [creator for creator in creators if creator.source_names == [source]]
        overlap = [creator for creator in source_creators if len(creator.source_names) > 1]
        email_records = sum(bool(record.email) for record in source_records)
        invalid_records = sum(bool(record.invalid_fields) for record in source_records)
        source_rows.append(
            [
                source,
                len(source_records),
                len(source_creators),
                len(source_unique),
                len(overlap),
                email_records,
                email_records / len(source_records) if source_records else 0,
                invalid_records,
            ]
        )
    source_end = write_rows(
        source_sheet,
        source_headers,
        source_rows,
        table_name="SourceComparison",
    )
    for cell in source_sheet[f"G2:G{source_end}"]:
        cell[0].number_format = "0.0%"

    coverage_sheet = workbook.create_sheet("Field Coverage")
    coverage_headers = [
        "field",
        "populated_creators",
        "missing_creators",
        "coverage_rate",
        "stale_or_unknown",
        "conflict_count",
    ]
    critical_fields = (
        "platform",
        "handle",
        "profile_url",
        "followers",
        "average_views",
        "engagement_rate",
        "country",
        "language",
        "email",
        "latest_post_at",
        "latest_sponsored_post_at",
    )
    coverage_rows = []
    for field_name in critical_fields:
        populated = sum(creator.selected.get(field_name) is not None for creator in creators)
        stale = sum(
            item.field == field_name for creator in creators for item in creator.stale_items
        )
        conflicts = sum(
            item.field == field_name for creator in creators for item in creator.conflicts
        )
        coverage_rows.append(
            [
                field_name,
                populated,
                len(creators) - populated,
                populated / len(creators) if creators else 0,
                stale,
                conflicts,
            ]
        )
    coverage_end = write_rows(
        coverage_sheet,
        coverage_headers,
        coverage_rows,
        table_name="FieldCoverage",
    )
    for cell in coverage_sheet[f"D2:D{coverage_end}"]:
        cell[0].number_format = "0.0%"

    conflict_sheet = workbook.create_sheet("Conflicts")
    conflict_headers = ["creator_id", "field", "level", "values", "sources", "reason"]
    conflict_rows = [
        [item.creator_id, item.field, item.level, item.values, item.sources, item.reason]
        for creator in creators
        for item in creator.conflicts
    ]
    conflict_end = write_rows(
        conflict_sheet,
        conflict_headers,
        conflict_rows,
        table_name="FieldConflicts",
    )
    if conflict_end >= 2:
        for row in range(2, conflict_end + 1):
            level = conflict_sheet.cell(row, 3).value
            color = RED if level == "High" else ORANGE if level == "Medium" else YELLOW
            conflict_sheet.cell(row, 3).fill = PatternFill("solid", fgColor=color)

    stale_sheet = workbook.create_sheet("Stale Data")
    stale_headers = [
        "creator_id",
        "field",
        "selected_value",
        "source",
        "observed_at",
        "ttl_days",
        "status",
    ]
    stale_rows = [
        [
            item.creator_id,
            item.field,
            stringify(item.selected_value),
            item.source,
            item.observed_at,
            item.ttl_days,
            item.status,
        ]
        for creator in creators
        for item in creator.stale_items
    ]
    stale_end = write_rows(
        stale_sheet,
        stale_headers,
        stale_rows,
        table_name="StaleData",
    )
    if stale_end >= 2:
        for row in range(2, stale_end + 1):
            stale_sheet.cell(row, 5).number_format = "yyyy-mm-dd hh:mm"

    duplicate_sheet = workbook.create_sheet("Duplicate Candidates")
    duplicate_headers = [
        "creator_id",
        "source_record_ids",
        "sources",
        "source_record_count",
        "merge_basis",
    ]
    duplicate_rows = [
        [
            creator.creator_id,
            ", ".join(creator.source_record_ids),
            ", ".join(creator.source_names),
            len(creator.records),
            creator.merge_basis,
        ]
        for creator in creators
        if len(creator.records) > 1
    ]
    write_rows(
        duplicate_sheet,
        duplicate_headers,
        duplicate_rows,
        table_name="DuplicateCandidates",
    )

    workbook.save(path)


def write_creator_shortlist(path: Path, creators: list[CreatorAggregate]) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Shortlist"
    headers = [
        "creator_id",
        "action_level",
        "reference_score",
        "platform",
        "handle",
        "profile_url",
        "followers",
        "average_views",
        "engagement_rate",
        "country",
        "language",
        "brand_fit",
        "commercial_readiness",
        "contactability",
        "data_confidence",
        "why_contact",
        "sources_used",
        "selected_email",
        "email_source",
        "email_observed_at",
        "data_warnings",
        "refresh_recommendation",
        "exclusion_reasons",
        "source_record_ids",
        "user_accepted",
        "user_note",
    ]
    rows = []
    for creator in creators:
        selected = creator.selected
        email_meta = creator.selected_meta.get("email", {})
        rows.append(
            [
                creator.creator_id,
                creator.action_level,
                creator.reference_score,
                selected.get("platform"),
                selected.get("handle"),
                selected.get("profile_url"),
                selected.get("followers"),
                selected.get("average_views"),
                selected.get("engagement_rate"),
                selected.get("country"),
                selected.get("language"),
                creator.dimensions["brand_fit"],
                creator.dimensions["commercial_readiness"],
                creator.dimensions["contactability"],
                creator.dimensions["data_confidence"],
                "\n".join(creator.why_contact),
                ", ".join(creator.source_names),
                selected.get("email"),
                email_meta.get("source", ""),
                email_meta.get("observed_at"),
                "\n".join(creator.warnings),
                ", ".join(creator.refresh_recommendation),
                ", ".join(creator.exclusion_reasons),
                ", ".join(creator.source_record_ids),
                "",
                "",
            ]
        )
    end_row = write_rows(
        worksheet,
        headers,
        rows,
        table_name="CreatorShortlist",
    )
    for row in range(2, end_row + 1):
        action = worksheet.cell(row, 2).value
        color = {"Priority": GREEN, "Verify": YELLOW, "Hold": ORANGE, "Excluded": RED}[action]
        worksheet.cell(row, 2).fill = PatternFill("solid", fgColor=color)
        worksheet.cell(row, 3).number_format = "0.00"
        worksheet.cell(row, 7).number_format = "#,##0"
        worksheet.cell(row, 8).number_format = "#,##0"
        worksheet.cell(row, 9).number_format = "0.0%"
        for column in range(12, 16):
            worksheet.cell(row, column).number_format = "0.00"
        worksheet.cell(row, 20).number_format = "yyyy-mm-dd hh:mm"
        worksheet.cell(row, 25).fill = PatternFill("solid", fgColor=YELLOW)
        worksheet.cell(row, 26).fill = PatternFill("solid", fgColor=YELLOW)
    validation = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
    worksheet.add_data_validation(validation)
    validation.add(f"Y2:Y{max(end_row, 2)}")
    workbook.save(path)


def write_merge_review(path: Path, review_items: list[dict[str, str]]) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Review Queue"
    headers = [
        "candidate_a",
        "candidate_b",
        "match_reason",
        "conflict_reason",
        "suggested_action",
        "user_decision",
        "user_note",
    ]
    rows = [[item.get(header, "") for header in headers] for item in review_items]
    end_row = write_rows(
        worksheet,
        headers,
        rows,
        table_name="MergeReviewQueue",
    )
    for row in range(2, end_row + 1):
        worksheet.cell(row, 6).fill = PatternFill("solid", fgColor=YELLOW)
        worksheet.cell(row, 7).fill = PatternFill("solid", fgColor=YELLOW)
    validation = DataValidation(
        type="list",
        formula1='"merge,keep_separate,unsure"',
        allow_blank=True,
    )
    worksheet.add_data_validation(validation)
    validation.add(f"F2:F{max(end_row, 2)}")
    workbook.save(path)


def write_feedback_template(path: Path, creators: list[CreatorAggregate]) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Feedback"
    headers = [
        "creator_id",
        "action_level",
        "reference_score",
        "platform",
        "handle",
        "sources_used",
        "merge_correct",
        "shortlist_accepted",
        "contact_correct",
        "actually_contacted",
        "delivered",
        "bounced",
        "replied",
        "positive_reply",
        "feedback_note",
    ]
    rows = [
        [
            creator.creator_id,
            creator.action_level,
            creator.reference_score,
            creator.selected.get("platform"),
            creator.selected.get("handle"),
            ", ".join(creator.source_names),
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ]
        for creator in creators
    ]
    end_row = write_rows(
        worksheet,
        headers,
        rows,
        table_name="CreatorFeedbackTemplate",
        freeze="G2",
    )
    for row in range(2, end_row + 1):
        worksheet.cell(row, 3).number_format = "0.00"
        for column in range(7, 16):
            worksheet.cell(row, column).fill = PatternFill("solid", fgColor=YELLOW)
    validation = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
    worksheet.add_data_validation(validation)
    for column in range(7, 15):
        validation.add(
            f"{get_column_letter(column)}2:{get_column_letter(column)}{max(end_row, 2)}"
        )
    worksheet.column_dimensions["O"].width = 48
    workbook.save(path)


def write_feedback_report(
    path: Path,
    *,
    run_id: str,
    metrics: list[dict[str, Any]],
    feedback_rows: list[dict[str, Any]],
    notes: list[dict[str, Any]],
    warnings: list[str],
) -> None:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    summary.sheet_view.showGridLines = False
    summary.merge_cells("A1:D1")
    summary["A1"] = "Creator Outreach Feedback Report"
    summary["A1"].fill = PatternFill("solid", fgColor=DARK_BLUE)
    summary["A1"].font = Font(name="Aptos Display", size=16, bold=True, color=WHITE)
    summary["A1"].alignment = Alignment(vertical="center")
    summary.row_dimensions[1].height = 34
    append_safe_row(summary, ["Run ID", run_id, "", ""])
    append_safe_row(summary, ["Metric", "Value", "Numerator", "Denominator"])
    for metric in metrics:
        append_safe_row(
            summary,
            [
                metric["label"],
                metric["value"],
                metric.get("numerator"),
                metric.get("denominator"),
            ],
        )
    summary.column_dimensions["A"].width = 28
    summary.column_dimensions["B"].width = 18
    summary.column_dimensions["C"].width = 14
    summary.column_dimensions["D"].width = 14
    summary.freeze_panes = "A4"
    for row in summary.iter_rows(min_row=2, max_row=summary.max_row):
        for cell in row:
            cell.font = Font(name="Aptos", size=10, color=TEXT)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=THIN_GRAY)
    style_header(summary, 3, 1, 4)
    for row_number, metric in enumerate(metrics, start=4):
        if metric["kind"] == "rate" and metric["value"] is not None:
            summary.cell(row_number, 2).number_format = "0.0%"
        elif metric["kind"] == "count":
            summary.cell(row_number, 2).number_format = "#,##0"
    add_table(summary, "FeedbackSummary", 3, summary.max_row, 4)

    feedback_sheet = workbook.create_sheet("Imported Feedback")
    feedback_headers = [
        "creator_id",
        "action_level",
        "reference_score",
        "platform",
        "handle",
        "merge_correct",
        "shortlist_accepted",
        "contact_correct",
        "actually_contacted",
        "delivered",
        "bounced",
        "replied",
        "positive_reply",
        "feedback_note",
    ]
    feedback_end = write_rows(
        feedback_sheet,
        feedback_headers,
        [[row.get(header) for header in feedback_headers] for row in feedback_rows],
        table_name="ImportedFeedback",
        freeze="F2",
    )
    for row in range(2, feedback_end + 1):
        feedback_sheet.cell(row, 3).number_format = "0.00"
    feedback_sheet.column_dimensions["N"].width = 48

    notes_sheet = workbook.create_sheet("Text Feedback")
    note_headers = ["feedback_note", "occurrences", "creator_ids"]
    write_rows(
        notes_sheet,
        note_headers,
        [
            [item["feedback_note"], item["occurrences"], ", ".join(item["creator_ids"])]
            for item in notes
        ],
        table_name="TextFeedbackSummary",
    )
    notes_sheet.column_dimensions["A"].width = 68
    notes_sheet.column_dimensions["C"].width = 42

    warning_sheet = workbook.create_sheet("Validation Warnings")
    write_rows(
        warning_sheet,
        ["warning"],
        [[warning] for warning in warnings] or [["No validation warnings."]],
        table_name="FeedbackWarnings",
    )
    warning_sheet.column_dimensions["A"].width = 100
    workbook.save(path)


def mask_email(email: str | None) -> str:
    if not email or "@" not in email:
        return ""
    local, domain = email.split("@", 1)
    visible = local[:2]
    return f"{visible}***@{domain}"


def markdown_escape(value: Any) -> str:
    return stringify(value).replace("|", "\\|").replace("\n", " ")


def write_report(
    path: Path,
    *,
    run_id: str,
    campaign: Campaign,
    campaign_parse_result: CampaignParseResult,
    records: list[NormalizedRecord],
    creators: list[CreatorAggregate],
    review_items: list[dict[str, str]],
    mapping_summaries: list[dict[str, Any]],
    reference_time: datetime,
    top: int,
    run_type: str = "standard",
    data_classification: str = "user_provided",
) -> None:
    action_counts = Counter(creator.action_level for creator in creators)
    conflict_count = sum(len(creator.conflicts) for creator in creators)
    high_conflicts = sum(
        item.level == "High" for creator in creators for item in creator.conflicts
    )
    stale_count = sum(len(creator.stale_items) for creator in creators)
    invalid_count = sum(len(record.invalid_fields) for record in records)
    unresolved_reviews = sum(
        item.get("user_decision") not in {"merge", "keep_separate"}
        for item in review_items
    )
    resolved_reviews = len(review_items) - unresolved_reviews
    recommended = [creator for creator in creators if creator.action_level in {"Priority", "Verify"}][:top]

    report_title = (
        "# Fully Synthetic Demo — Creator Signal Intelligence Audit Report"
        if run_type == "demo"
        else "# Creator Signal Intelligence — Audit Report"
    )
    lines = [
        report_title,
        "",
        f"> Data classification: `{data_classification}`  ",
        f"> Run ID: `{run_id}`  ",
        f"> Reference time: `{reference_time.isoformat()}`  ",
        "> Boundary: local files only; no model, API, crawling, email sending, or external write.",
        "",
        "## Campaign 摘要",
        "",
        f"- 品牌：{campaign.brand}",
        f"- 目标市场：{', '.join(campaign.markets) or '未设置'}",
        f"- 内容语言：{', '.join(campaign.languages) or '未设置'}",
        f"- 平台：{', '.join(campaign.platforms) or '未设置'}",
        f"- 粉丝范围：{campaign.follower_min or '未设置'}–{campaign.follower_max or '未设置'}",
        f"- 最近发帖：{campaign.latest_post_max_age_days or '未设置'} 天内",
        (
            f"- 必须有联系路径：{'是' if campaign.require_contact_path else '否'}"
            if campaign.require_contact_path is not None
            else "- 必须有联系路径：未设置"
        ),
        f"- Blocklist：{', '.join(campaign.blocklist) or '未设置'}",
        f"- Topic 标签：{', '.join(campaign.topics) or '未设置'}",
        "",
        "## Campaign 条件解析",
        "",
        f"- Parser Version：{campaign_parse_result.parser_version}",
        "",
        "### 已识别条件",
        "",
    ]
    if campaign_parse_result.recognized_conditions:
        for condition in campaign_parse_result.recognized_conditions:
            applied = "参与本轮规则" if condition.get("applied") else "仅保留，未参与规则"
            lines.append(
                f"- `{condition['field']}`：{markdown_escape(condition['value'])}（{applied}）"
            )
    else:
        lines.append("- 无。")

    lines.extend(["", "### 未识别条件", ""])
    if campaign_parse_result.unrecognized_conditions:
        for condition in campaign_parse_result.unrecognized_conditions:
            lines.append(
                f"- [{condition['severity']}] {markdown_escape(condition['text'])}"
            )
    else:
        lines.append("- 无。")

    lines.extend(["", "### 冲突条件", ""])
    if campaign_parse_result.conflicting_conditions:
        for condition in campaign_parse_result.conflicting_conditions:
            lines.append(
                f"- `{condition['field']}`：{markdown_escape(condition['reason'])}"
            )
    else:
        lines.append("- 无。")

    lines.extend(["", "### 解析警告", ""])
    if campaign_parse_result.warnings:
        lines.extend(f"- {markdown_escape(warning)}" for warning in campaign_parse_result.warnings)
    else:
        lines.append("- 无。")

    lines.extend(
        [
        "",
        "## 数据审计摘要",
        "",
        f"- 输入文件：{len(mapping_summaries)}",
        f"- 原始记录：{len(records)}",
        f"- 去重后达人：{len(creators)}",
        f"- 确定性去重减少：{len(records) - len(creators)} 条记录",
        f"- 待人工审核：{unresolved_reviews} 项",
        f"- 已处理审核：{resolved_reviews} 项",
        f"- 字段冲突：{conflict_count} 项，其中 High {high_conflicts} 项",
        f"- 过期或观测时间未知：{stale_count} 项",
        f"- 无效原始值：{invalid_count} 项",
        "",
        "## 行动等级",
        "",
        "| Level | Count |",
        "|---|---:|",
        ]
    )
    for level in ("Priority", "Verify", "Hold", "Excluded"):
        lines.append(f"| {level} | {action_counts.get(level, 0)} |")

    lines.extend(
        [
            "",
            f"## Top {min(top, len(recommended))} 候选概览",
            "",
            "| Rank | Level | Creator | Score | Sources | Email | Why | Warnings |",
            "|---:|---|---|---:|---|---|---|---|",
        ]
    )
    for rank, creator in enumerate(recommended, start=1):
        selected = creator.selected
        creator_label = f"{selected.get('platform') or ''}/@{selected.get('handle') or ''}"
        lines.append(
            "| "
            + " | ".join(
                [
                    str(rank),
                    creator.action_level,
                    markdown_escape(creator_label),
                    f"{creator.reference_score:.2f}",
                    markdown_escape(", ".join(creator.source_names)),
                    markdown_escape(mask_email(selected.get("email"))),
                    markdown_escape("；".join(creator.why_contact[:2])),
                    markdown_escape("；".join(creator.warnings[:2]) or "无高风险提醒"),
                ]
            )
            + " |"
        )

    lines.extend(
        [
            "",
            "## Mapping 结果",
            "",
            "| File | Source | Mapping | Validation | Rows | Mapped columns | Unmapped columns |",
            "|---|---|---|---|---:|---:|---|",
        ]
    )
    for item in mapping_summaries:
        lines.append(
            f"| {markdown_escape(Path(item['file']).name)} | {item['source']} | "
            f"{item.get('mapping_origin', 'manual')} | "
            f"{item.get('adapter_validation_level') or 'Generic Import'} | "
            f"{item['row_count']} | "
            f"{item['mapped_columns']} | "
            f"{markdown_escape(', '.join(item['unmapped_columns']))} |"
        )

    lines.extend(
        [
            "",
            "## 使用限制",
            "",
            "- Native Adapter 使用确定性映射，但验证等级必须单独查看；命中 Adapter 不等于格式已 Verified。",
            "- Generic Mapping 使用列名和样例值生成建议。",
            "- 歧义字段必须经用户确认或复用已确认配置，不会由系统静默决定。",
            "- Brand Fit 在没有结构化内容标签时只使用市场与语言证据，不分析达人近期内容。",
            "- 评分是 Open Alpha 初始假设，不代表真实合作概率。",
            "- 跨平台同 Handle、共享经纪邮箱和矛盾身份键不会自动合并。",
            "- 本报告只反映本次输入快照；过期或观测时间未知的数据需要人工补查。",
            "",
            "## 建议下一步",
            "",
            (
                "1. 先处理 `merge_review.xlsx` 中尚未确认的身份疑点。"
                if unresolved_reviews
                else "1. 身份审核已完成；如无新增证据，无需重复处理 Review Queue。"
            ),
            "2. 对 Verify 候选补查冲突、过期或缺失字段。",
            "3. 在 `creator_shortlist.xlsx` 中人工确认 Priority 候选后再进入现有工作表。",
        ]
    )
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")
